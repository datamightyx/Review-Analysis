"""
Pull FBA customer returns straight from SP-API and build the same
mismatch-analysis Excel report as pages/1_Returns_Analysis.py.

Fetches GET_FBA_FULFILLMENT_CUSTOMER_RETURNS_DATA (tab-delimited: return-date,
order-id, sku, asin, fnsku, product-name, quantity, fulfillment-center-id,
detailed-disposition, reason, status, license-plate-number, customer-comments),
then runs the comment-vs-stated-reason mismatch analysis and writes an .xlsx
with a SUMMARY sheet plus one sheet per ASIN.

Auth: raw SP-API REST calls (LWA bearer token only). Credentials read from
.env via python-dotenv (same vars as sp_api_refunds_check.py).

Usage:
    python returns_analysis_api.py --start 2026-05-01 --end 2026-08-14
    python returns_analysis_api.py --start 2026-05-01 --end 2026-08-14 --output MyReport.xlsx
"""

import argparse
import csv
import gzip
import io
import os
import re
import sys
import time

import pandas as pd
import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

LWA_TOKEN_URL = "https://api.amazon.com/auth/o2/token"
REPORT_TYPE = "GET_FBA_FULFILLMENT_CUSTOMER_RETURNS_DATA"

REQUIRED_ENV = [
    "SP_API_LWA_CLIENT_ID",
    "SP_API_LWA_CLIENT_SECRET",
    "SP_API_REFRESH_TOKEN",
    "SP_API_MARKETPLACE_ID",
    "SP_API_ENDPOINT",
]


# ── SP-API fetch (same pattern as sp_api_refunds_check.py) ─────────────────

def get_config():
    missing = [k for k in REQUIRED_ENV if not os.environ.get(k)]
    if missing:
        sys.exit(f"Missing .env values: {', '.join(missing)}")
    return {k: os.environ[k] for k in REQUIRED_ENV}


def get_access_token(cfg):
    resp = requests.post(
        LWA_TOKEN_URL,
        data={
            "grant_type": "refresh_token",
            "refresh_token": cfg["SP_API_REFRESH_TOKEN"],
            "client_id": cfg["SP_API_LWA_CLIENT_ID"],
            "client_secret": cfg["SP_API_LWA_CLIENT_SECRET"],
        },
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()["access_token"]


def create_report(cfg, access_token, start, end):
    url = f"{cfg['SP_API_ENDPOINT']}/reports/2021-06-30/reports"
    body = {
        "reportType": REPORT_TYPE,
        "marketplaceIds": [cfg["SP_API_MARKETPLACE_ID"]],
        "dataStartTime": f"{start}T00:00:00Z",
        "dataEndTime": f"{end}T23:59:59Z",
    }
    resp = requests.post(url, json=body, headers={"x-amz-access-token": access_token}, timeout=30)
    resp.raise_for_status()
    return resp.json()["reportId"]


def poll_report(cfg, access_token, report_id, interval=20, timeout=1800):
    url = f"{cfg['SP_API_ENDPOINT']}/reports/2021-06-30/reports/{report_id}"
    waited = 0
    while waited <= timeout:
        resp = requests.get(url, headers={"x-amz-access-token": access_token}, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        status = data["processingStatus"]
        print(f"  report {report_id}: {status} ({waited}s elapsed)")
        if status == "DONE":
            return data["reportDocumentId"]
        if status in ("FATAL", "CANCELLED"):
            sys.exit(f"Report {report_id} failed: {status}")
        time.sleep(interval)
        waited += interval
    sys.exit(f"Report {report_id} did not finish within {timeout}s")


def download_report_document(cfg, access_token, document_id):
    url = f"{cfg['SP_API_ENDPOINT']}/reports/2021-06-30/documents/{document_id}"
    resp = requests.get(url, headers={"x-amz-access-token": access_token}, timeout=30)
    resp.raise_for_status()
    doc = resp.json()

    file_resp = requests.get(doc["url"], timeout=120)
    file_resp.raise_for_status()
    raw = file_resp.content

    if doc.get("compressionAlgorithm") == "GZIP":
        raw = gzip.decompress(raw)

    try:
        return raw.decode("utf-8")
    except UnicodeDecodeError:
        return raw.decode("cp1252")


def fetch_customer_returns(cfg, start, end):
    print("Requesting LWA access token...")
    access_token = get_access_token(cfg)

    print(f"Creating report {REPORT_TYPE} for {start}..{end}...")
    report_id = create_report(cfg, access_token, start, end)

    print("Polling report status...")
    document_id = poll_report(cfg, access_token, report_id)

    print("Downloading report document...")
    return download_report_document(cfg, access_token, document_id)


def parse_customer_returns(raw_tsv_text):
    # QUOTE_NONE: customer-comments is unquoted free text that may start with a
    # quote char, which the default parser would treat as a quoted field and
    # swallow following lines into one row
    df = pd.read_csv(io.StringIO(raw_tsv_text), sep="\t", dtype=str, quoting=csv.QUOTE_NONE)
    return normalize_returns_frame(df)


def normalize_returns_frame(df):
    df = df.copy()
    df.columns = df.columns.str.strip().str.lower()

    required = ["asin", "product-name", "reason", "customer-comments"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        sys.exit(f"Report missing expected columns {missing}. Got: {list(df.columns)}")

    if "return-date" in df.columns:
        df["return-date"] = pd.to_datetime(df["return-date"], errors="coerce", utc=True).dt.tz_localize(None)
    else:
        df["return-date"] = pd.NaT

    if "quantity" in df.columns:
        df["quantity"] = pd.to_numeric(df["quantity"], errors="coerce")
    else:
        df["quantity"] = None

    return df


# ── Refund-vs-return reconciliation ────────────────────────────────────────
#
# The customer-returns report only contains units that physically came back and
# were scanned at an FC. Business Reports "Refunds" (and Refund rows in the
# Unified Transaction report) count money refunded, which also covers returnless
# refunds, items the customer never shipped back, and returns still in transit.
# Refunds are therefore always >= returns. Passing --transactions appends one
# row per refund that has no matching return so the report totals tie out to
# the refund count; those rows carry no customer comment by definition.

REFUND_NO_RETURN = "REFUND_NO_RETURN"
STATUS_REFUND_ONLY = "Refund - No Return"


FINANCES_PATH = "/finances/v0/financialEvents"


def fetch_refund_events(cfg, start, end, release_lag_days=7, progress=None):
    """Refund events for the window straight from the Finances API.

    GET_DATE_RANGE_FINANCIAL_TRANSACTION_DATA (report type 1202) is not grantable
    for this account ("Request for report type 1202 is not allowed at this
    time"), so refunds come from /finances/v0/financialEvents instead. Windows
    are UTC, matching the returns report; the Unified Transaction CSV is PT.

    PostedDate here is the *release* timestamp, not the moment of the refund:
    verified against the Unified Transaction CSV, every late row lines up with
    its "Transaction Release Date" (refund Aug 10 -> API 2026-08-16T14:51:40Z),
    and refunds still Deferred are absent entirely. Release runs a few days
    behind, so the upper bound is padded by release_lag_days to catch refunds
    made inside the window but released after it. The trade-off: refunds truly
    made in those extra days are pulled in too. Pass 0 for a strict window.

    Returns the normalized frame: order_id, sku, quantity, posted (naive UTC).
    """
    access_token = get_access_token(cfg)
    url = f"{cfg['SP_API_ENDPOINT']}{FINANCES_PATH}"
    headers = {"x-amz-access-token": access_token}

    upper = pd.Timestamp(end) + pd.Timedelta(days=release_lag_days, hours=23, minutes=59)
    # the API rejects a PostedBefore in the future
    now_utc = pd.Timestamp.utcnow().tz_localize(None) - pd.Timedelta(minutes=5)
    upper = min(upper, now_utc)

    rows, next_token, page = [], None, 0
    while True:
        if next_token:
            params = {"NextToken": next_token}
        else:
            params = {
                "MaxResultsPerPage": 100,
                "PostedAfter": f"{start}T00:00:00Z",
                "PostedBefore": upper.strftime("%Y-%m-%dT%H:%M:%SZ"),
            }
        resp = requests.get(url, params=params, headers=headers, timeout=60)
        if resp.status_code == 429:
            time.sleep(5)
            continue
        resp.raise_for_status()
        payload = resp.json()["payload"]

        for event in payload.get("FinancialEvents", {}).get("RefundEventList", []):
            for item in event.get("ShipmentItemAdjustmentList") or []:
                rows.append({
                    "order_id": event.get("AmazonOrderId"),
                    "sku":      item.get("SellerSKU"),
                    "quantity": item.get("QuantityShipped"),
                    "posted":   event.get("PostedDate"),
                })

        page += 1
        if progress:
            progress(page, len(rows))
        next_token = payload.get("NextToken")
        if not next_token:
            break
        time.sleep(2.2)  # financialEvents is rate limited to 0.5 requests/sec

    ref = pd.DataFrame(rows, columns=["order_id", "sku", "quantity", "posted"])
    ref["quantity"] = pd.to_numeric(ref["quantity"], errors="coerce").fillna(0)
    ref = ref[ref["quantity"] > 0]
    ref["posted"] = pd.to_datetime(ref["posted"], errors="coerce", utc=True).dt.tz_localize(None)
    return ref.reset_index(drop=True)


def parse_transaction_refunds(transactions_path):
    """Refund rows from a Unified Transaction CSV, in the same normalized shape."""
    with open(transactions_path, "rb") as f:
        raw = f.read()
    for enc in ("utf-8", "utf-8-sig", "cp1252", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw.decode("utf-8", errors="replace")

    # slice on the raw text, not splitlines(): data rows contain quoted fields
    # with embedded newlines, so line-splitting would corrupt them
    m = re.search(r'^"?date/time"?\s*,', text, re.IGNORECASE | re.MULTILINE)
    if m is None:
        sys.exit(f"{transactions_path}: could not find a 'date/time' header row")

    tx = pd.read_csv(io.StringIO(text[m.start():]), dtype=str)
    tx.columns = [c.strip() for c in tx.columns]
    tx["quantity"] = pd.to_numeric(tx["quantity"], errors="coerce").fillna(0)

    ref = tx[tx["type"].astype(str).str.strip().str.lower() == "refund"].copy()
    ref = ref[ref["quantity"] > 0]  # quantity 0 == fee/tax-only adjustment line
    posted = pd.to_datetime(
        ref["date/time"].str.replace(r"\s+P[DS]T$", "", regex=True),
        format="%b %d, %Y %I:%M:%S %p",
        errors="coerce",
    )
    return pd.DataFrame({
        "order_id": ref["order id"].values,
        "sku":      ref["sku"].values,
        "quantity": ref["quantity"].values,
        "posted":   posted.values,
    })


def refunds_to_return_rows(ref, returns_df, verbose=True):
    """Refund events with no matching return line, shaped like returns-report rows.

    Match key is (order id, SKU). SKUs that never appear in the returns report
    cannot be mapped to an ASIN, so they are dropped and reported.
    """
    sku_asin = (
        returns_df.dropna(subset=["sku", "asin"])
        .groupby("sku")["asin"].agg(lambda s: s.value_counts().index[0]).to_dict()
    )
    sku_name = (
        returns_df.dropna(subset=["sku", "product-name"])
        .groupby("sku")["product-name"].first().to_dict()
    )

    ret_keys = set(zip(returns_df["order-id"], returns_df["sku"]))
    ref = ref[[(o, s) not in ret_keys for o, s in zip(ref["order_id"], ref["sku"])]]

    skipped = sorted(ref[~ref["sku"].isin(sku_asin)]["sku"].dropna().unique().tolist())
    if skipped and verbose:
        print(f"  note: refund SKUs with no ASIN in the returns report, skipped: {skipped}")
    ref = ref[ref["sku"].isin(sku_asin)]

    rows = pd.DataFrame({
        "return-date":        ref["posted"].values,
        "order-id":           ref["order_id"].values,
        "sku":                ref["sku"].values,
        "asin":               ref["sku"].map(sku_asin).values,
        "product-name":       ref["sku"].map(sku_name).values,
        "quantity":           ref["quantity"].values,
        "reason":             REFUND_NO_RETURN,
        "customer-comments":  pd.NA,
    })
    return rows, skipped


def load_refund_only_rows(transactions_path, returns_df):
    """CLI path: Unified Transaction CSV -> refund rows with no matching return."""
    ref = parse_transaction_refunds(transactions_path)
    rows, _ = refunds_to_return_rows(ref, returns_df)
    return rows


# ── Mismatch analysis (same logic as pages/1_Returns_Analysis.py) ──────────

KW = {
    'SIZE_TOO_LARGE': [
        'too large','too big','too wide','too long','too tall','runs large',
        'incorrect size','wrong size','size is wrong','too bulky','size too big',
        'bigger than','larger than','size larger','bit too big',
    ],
    'SIZE_TOO_SMALL': [
        'too small','too tight','too narrow','too short','runs small','size too small',
        'smaller than','too tiny','not big enough','too little','size smaller',
    ],
    'DEFECTIVE': [
        'defective','broken',"doesn't work",'does not work','not work','broke',
        'malfunction','stopped working','falling off','falls off','cracked',
        'leaking',' leak ','torn','ripped','damaged','doesnt function',
        "doesn't function",'no longer work',
    ],
    'QUALITY_ISSUE': [
        'poor quality','bad quality','too thin','very thin','flimsy','not durable',
        'low quality','inferior','terrible quality','bad material',
        'quality is bad','not good quality','not the same quality','cheap quality',
        'quality is poor','material is thin',
    ],
    'NOT_AS_DESCRIBED': [
        'not as described','not as expected','not what i expected',
        'misleading','misrepresented','inaccurate description',
    ],
    'WRONG_ITEM': [
        'wrong item','wrong product','wrong color','ordered wrong',
    ],
    'DELIVERY_ISSUE': [
        'never arrived','did not arrive','not delivered','never received',
        'late delivery','delayed','shipping was','weeks delayed','not arrive',
        'took too long','not received',
    ],
    'CHANGED_MIND': [
        'changed mind',"don't need",'do not need','no longer need',
        'no longer want','changed my mind',"don't want",'dont need','dont want',
    ],
    'BETTER_PRICE': [
        'found cheaper','better price','found better price','cheaper elsewhere',
        'cheaper on','lower price',
    ],
    'SIZE_ISSUE': [
        ' size ',' fit ',' fits ','fitting','incorrect size','wrong size',
        "doesn't fit","does not fit","wont fit","won't fit",
    ],
}

REASON_OK = {
    'APPAREL_TOO_LARGE':         ['SIZE_TOO_LARGE','SIZE_ISSUE'],
    'APPAREL_TOO_SMALL':         ['SIZE_TOO_SMALL','SIZE_ISSUE'],
    'POOR_FIT':                  ['SIZE_TOO_LARGE','SIZE_TOO_SMALL','SIZE_ISSUE'],
    'DEFECTIVE':                 ['DEFECTIVE'],
    'NOT_AS_DESCRIBED':          ['NOT_AS_DESCRIBED','QUALITY_ISSUE','SIZE_TOO_LARGE','SIZE_TOO_SMALL'],
    'QUALITY_UNACCEPTABLE':      ['QUALITY_ISSUE','DEFECTIVE'],
    'ORDERED_WRONG_ITEM':        ['WRONG_ITEM'],
    'UNWANTED_ITEM':             ['CHANGED_MIND','BETTER_PRICE'],
    'FOUND_BETTER_PRICE':        ['BETTER_PRICE','CHANGED_MIND'],
    'MISSING_PARTS':             ['DEFECTIVE'],
    'DAMAGED_BY_FC':             ['DEFECTIVE'],
    'DAMAGED_BY_CARRIER':        ['DEFECTIVE'],
    'NEVER_ARRIVED':             ['DELIVERY_ISSUE'],
    'MISSED_ESTIMATED_DELIVERY': ['DELIVERY_ISSUE'],
    'UNDELIVERABLE_UNKNOWN':     ['DELIVERY_ISSUE'],
    'UNDELIVERABLE_REFUSED':     ['DELIVERY_ISSUE'],
    'NO_REASON_GIVEN':           [],
    'SWITCHEROO':                ['WRONG_ITEM'],
    'NOT_COMPATIBLE':            ['NOT_AS_DESCRIBED','SIZE_ISSUE'],
    'PART_NOT_COMPATIBLE':       ['NOT_AS_DESCRIBED','SIZE_ISSUE'],
    'MISORDERED':                ['WRONG_ITEM'],
    'EXTRA_ITEM':                [],
    'EXCESSIVE_INSTALLATION':    [],
    'UNAUTHORIZED_PURCHASE':     [],
}

TRUE_REASON_LABEL = {
    'SIZE_TOO_LARGE':   'SIZE - Too Large',
    'SIZE_TOO_SMALL':   'SIZE - Too Small',
    'SIZE_ISSUE':       'SIZE - Fit Issue',
    'DEFECTIVE':        'DEFECTIVE',
    'QUALITY_ISSUE':    'QUALITY ISSUE',
    'NOT_AS_DESCRIBED': 'NOT AS DESCRIBED',
    'WRONG_ITEM':       'WRONG ITEM',
    'DELIVERY_ISSUE':   'DELIVERY ISSUE',
    'CHANGED_MIND':     'CHANGED MIND',
    'BETTER_PRICE':     'FOUND BETTER PRICE',
}

TRIVIAL = {'', 'na', 'n/a', 'no', 'yes', 'z', 'return', 'ok', 'none', '.', '-', 'n'}

TOPIC_DISPLAY = {
    'SIZE_TOO_LARGE':   'SIZE — Too Large',
    'SIZE_TOO_SMALL':   'SIZE — Too Small',
    'SIZE_ISSUE':       'SIZE — Fit Issue',
    'DEFECTIVE':        'Defective / Not Working',
    'QUALITY_ISSUE':    'Quality Issue',
    'NOT_AS_DESCRIBED': 'Not As Described',
    'WRONG_ITEM':       'Wrong Item Sent',
    'DELIVERY_ISSUE':   'Delivery Issue',
    'CHANGED_MIND':     'Changed Mind',
    'BETTER_PRICE':     'Found Better Price',
    'OTHER':            'Other / Unclear',
}

TOPIC_ROW_COLOR = {
    'SIZE_TOO_LARGE':   'FDE9D9',
    'SIZE_TOO_SMALL':   'FDE9D9',
    'SIZE_ISSUE':       'FDE9D9',
    'DEFECTIVE':        'FFD7D7',
    'QUALITY_ISSUE':    'EAD1DC',
    'NOT_AS_DESCRIBED': 'D9E1F2',
    'WRONG_ITEM':       'FCE4D6',
    'DELIVERY_ISSUE':   'DDEBF7',
    'CHANGED_MIND':     'E2EFDA',
    'BETTER_PRICE':     'D9F2F2',
    'OTHER':            'F2F2F2',
}

C_HEADER    = '1F3864'
C_HEADER_FG = 'FFFFFF'
C_MISMATCH  = 'FFD7D7'
C_MATCH     = 'D7F0D7'
C_UNCLEAR   = 'FFF3CC'
C_NOCOMMENT = 'F2F2F2'
C_SUMMARY_H = '2E75B6'
C_ALT_ROW   = 'EEF3FA'

C_REFUNDONLY = 'E4DFEC'

STATUS_FILL = {
    'Mismatch':         PatternFill('solid', fgColor=C_MISMATCH),
    'Match':            PatternFill('solid', fgColor=C_MATCH),
    'Unclear':          PatternFill('solid', fgColor=C_UNCLEAR),
    'No Comment':       PatternFill('solid', fgColor=C_NOCOMMENT),
    STATUS_REFUND_ONLY: PatternFill('solid', fgColor=C_REFUNDONLY),
}

body_font    = Font(name='Calibri', size=9)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=False)
wrap_align   = Alignment(horizontal='left',   vertical='top',    wrap_text=True)
left_align   = Alignment(horizontal='left',   vertical='center')

thin   = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def classify_comment(comment):
    if pd.isna(comment):
        return 'NO_COMMENT'
    c = str(comment).strip().lower()
    if c in TRIVIAL or len(c) <= 2:
        return 'NO_COMMENT'
    for cat in [
        'SIZE_TOO_LARGE','SIZE_TOO_SMALL','DEFECTIVE','QUALITY_ISSUE',
        'NOT_AS_DESCRIBED','WRONG_ITEM','DELIVERY_ISSUE',
        'CHANGED_MIND','BETTER_PRICE','SIZE_ISSUE',
    ]:
        if any(k in c for k in KW[cat]):
            return cat
    return 'OTHER'


def get_status(row):
    topic = row['comment_topic']
    if row['reason'] == REFUND_NO_RETURN:
        return STATUS_REFUND_ONLY
    if topic == 'NO_COMMENT':
        return 'No Comment'
    if topic == 'OTHER':
        return 'Unclear'
    expected = REASON_OK.get(str(row['reason']), [])
    return 'Match' if topic in expected else 'Mismatch'


def get_true_reason(row):
    if row['status'] == 'Mismatch':
        return TRUE_REASON_LABEL.get(row['comment_topic'], row['reason'])
    return row['reason']


def pct_bar(pct, width=8):
    filled = max(0, min(width, round(pct / 100 * width)))
    return '█' * filled + '░' * (width - filled)


def style_header_row(ws, row_num, col_count, bg_color=C_HEADER):
    fill = PatternFill('solid', fgColor=bg_color)
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill = fill
        cell.font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
        cell.alignment = center_align
        cell.border = border
    ws.row_dimensions[row_num].height = 18


def write_analysis_block(ws, grp, start_row):
    n_total     = len(grp)
    n_noc       = (grp['status'] == 'No Comment').sum()
    n_refonly   = (grp['status'] == STATUS_REFUND_ONLY).sum()
    n_commented = n_total - n_noc - n_refonly
    n_mismatch  = (grp['status'] == 'Mismatch').sum()

    topic_vc = (
        grp[grp['comment_topic'] != 'NO_COMMENT']['comment_topic']
        .value_counts()
    )
    true_vc = grp['true_reason'].value_counts()

    mis_grp = grp[grp['status'] == 'Mismatch']
    mismatch_rows = []
    if len(mis_grp) > 0:
        for stated, sub in mis_grp.groupby('reason'):
            mismatch_rows.append({
                'stated':   stated,
                'count':    len(sub),
                'top_true': sub['true_reason'].value_counts().index[0],
            })
        mismatch_rows.sort(key=lambda x: x['count'], reverse=True)

    n_data = min(max(len(topic_vc), len(true_vc), len(mismatch_rows), 1), 12)

    T1_COLOR = '375623'
    T2_COLOR = '1F3864'
    T3_COLOR = '7B2D00'

    def hfill(c):
        return PatternFill('solid', fgColor=c)

    def hdr_font(c='FFFFFF', sz=9, bold=True):
        return Font(name='Calibri', bold=bold, size=sz, color=c)

    def brd():
        s = Side(style='thin', color='CCCCCC')
        return Border(left=s, right=s, top=s, bottom=s)

    la = Alignment(horizontal='left',   vertical='center')
    ca = Alignment(horizontal='center', vertical='center')

    def wc(row, col, val='', fill=None, font=None, align=None):
        cell = ws.cell(row, col)
        cell.value = val
        if fill:  cell.fill      = fill
        if font:  cell.font      = font
        if align: cell.alignment = align
        cell.border = brd()
        return cell

    r = start_row

    for (c1, c2), color, label in [
        ((1, 3), T1_COLOR, 'COMMENT TOPICS  —  what customers actually write'),
        ((5, 7), T2_COLOR, 'TRUE RETURN REASONS  —  after mismatch correction'),
        ((9,11), T3_COLOR, 'MISMATCH ANALYSIS  —  stated reason vs. comment'),
    ]:
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        wc(r, c1, label,
           fill=hfill(color),
           font=Font(name='Calibri', bold=True, size=10, color='FFFFFF'),
           align=Alignment(horizontal='left', vertical='center', indent=1))
    ws.row_dimensions[r].height = 18
    r += 1

    sub_defs = [
        (1,  'Comment Topic',           T1_COLOR),
        (2,  'Count',                   T1_COLOR),
        (3,  '% / bar',                 T1_COLOR),
        (5,  'True Reason',             T2_COLOR),
        (6,  'Count',                   T2_COLOR),
        (7,  '% of Total',              T2_COLOR),
        (9,  'Stated Reason (Amazon)',  T3_COLOR),
        (10, '# Mismatches',            T3_COLOR),
        (11, 'Corrected To',            T3_COLOR),
    ]
    for col, label, color in sub_defs:
        wc(r, col, label, fill=hfill(color), font=hdr_font(sz=8), align=ca)
    ws.row_dimensions[r].height = 14
    r += 1

    t1_items = list(topic_vc.items())
    t2_items = list(true_vc.items())

    for i in range(n_data):
        alt    = hfill('F7F7F7') if i % 2 else hfill('FFFFFF')
        df_fnt = Font(name='Calibri', size=9)
        sm_fnt = Font(name='Calibri', size=8)

        if i < len(t1_items):
            topic, cnt = t1_items[i]
            pct = cnt / n_commented * 100 if n_commented else 0
            rf  = hfill(TOPIC_ROW_COLOR.get(topic, 'F2F2F2'))
            wc(r, 1, TOPIC_DISPLAY.get(topic, topic), fill=rf, font=df_fnt, align=la)
            wc(r, 2, cnt,                              fill=rf, font=df_fnt, align=ca)
            wc(r, 3, f'{pct:.1f}%  {pct_bar(pct)}',   fill=rf, font=sm_fnt, align=la)
        else:
            for c in (1, 2, 3):
                wc(r, c, fill=hfill('FAFAFA'))

        if i < len(t2_items):
            reason, cnt = t2_items[i]
            pct = cnt / n_total * 100 if n_total else 0
            wc(r, 5, reason,                         fill=alt, font=df_fnt, align=la)
            wc(r, 6, cnt,                            fill=alt, font=df_fnt, align=ca)
            wc(r, 7, f'{pct:.1f}%  {pct_bar(pct)}', fill=alt, font=sm_fnt, align=la)
        else:
            for c in (5, 6, 7):
                wc(r, c, fill=hfill('FAFAFA'))

        if i < len(mismatch_rows):
            row_d    = mismatch_rows[i]
            cnt      = row_d['count']
            top_true = row_d['top_true']
            pct_m    = cnt / n_mismatch * 100 if n_mismatch else 0
            wc(r,  9, row_d['stated'],               fill=alt, font=df_fnt, align=la)
            wc(r, 10, f"{cnt}  ({pct_m:.0f}%)",      fill=alt, font=df_fnt, align=ca)
            wc(r, 11, top_true,
               fill=hfill('FFD7D7'),
               font=Font(name='Calibri', size=9, bold=True, color='7B2D00'),
               align=la)
        else:
            for c in (9, 10, 11):
                wc(r, c, fill=hfill('FAFAFA'))

        ws.row_dimensions[r].height = 14
        r += 1

    tf   = hfill('ECECEC')
    tfnt = Font(name='Calibri', bold=True, size=9)
    sfnt = Font(name='Calibri', bold=True, size=8)

    n_with_topic = int(topic_vc.sum()) if len(topic_vc) else 0
    pct_topic    = n_with_topic / n_commented * 100 if n_commented else 0
    wc(r, 1, 'TOTAL (with topic)',                       fill=tf, font=tfnt, align=la)
    wc(r, 2, n_with_topic,                               fill=tf, font=tfnt, align=ca)
    wc(r, 3, f'{pct_topic:.1f}% of {n_commented} commented', fill=tf, font=sfnt, align=la)

    wc(r, 5, 'TOTAL',                                    fill=tf, font=tfnt, align=la)
    wc(r, 6, int(true_vc.sum()) if len(true_vc) else 0,  fill=tf, font=tfnt, align=ca)
    wc(r, 7, '100%',                                     fill=tf, font=sfnt, align=la)

    mis_rate = n_mismatch / n_commented * 100 if n_commented else 0
    wc(r,  9, 'TOTAL MISMATCHES',                        fill=tf, font=tfnt, align=la)
    wc(r, 10, n_mismatch,                                fill=tf, font=tfnt, align=ca)
    wc(r, 11, f'{mis_rate:.1f}% of commented returns',   fill=tf, font=sfnt, align=la)

    ws.row_dimensions[r].height = 14
    r += 1
    return r


def build_workbook(df):
    df = df.copy()
    df['comment_topic'] = df['customer-comments'].apply(classify_comment)
    df['status']        = df.apply(get_status, axis=1)
    df['true_reason']   = df.apply(get_true_reason, axis=1)

    wb = Workbook()

    summary_cols = [
        'ASIN', 'Product Name', 'Total Cases (Returns + Refunds)', 'Physical Returns',
        'Returns with Comment',
        'Match', 'Mismatch', 'Unclear', 'No Comment', 'Refund w/o Return',
        'Mismatch %',
        'Top True Reason (Mismatch)',
        '2nd True Reason (Mismatch)',
        '3rd True Reason (Mismatch)',
    ]

    ws_sum = wb.active
    ws_sum.title = 'SUMMARY'
    ws_sum.append(summary_cols)
    style_header_row(ws_sum, 1, len(summary_cols), C_SUMMARY_H)

    for asin, grp in df.groupby('asin', sort=False):
        pname       = grp['product-name'].iloc[0]
        pname_short = pname[:80] + '...' if len(pname) > 80 else pname
        n_total     = len(grp)
        n_refonly   = (grp['status'] == STATUS_REFUND_ONLY).sum()
        n_returns   = n_total - n_refonly
        n_noc       = (grp['status'] == 'No Comment').sum()
        n_comment   = n_returns - n_noc
        n_match     = (grp['status'] == 'Match').sum()
        n_mis       = (grp['status'] == 'Mismatch').sum()
        n_unc       = (grp['status'] == 'Unclear').sum()
        mis_pct     = round(n_mis / n_comment * 100, 1) if n_comment > 0 else 0

        top_reasons = (
            grp[grp['status'] == 'Mismatch']['true_reason']
            .value_counts().head(3).index.tolist()
        )
        top_reasons += [''] * (3 - len(top_reasons))

        ws_sum.append([
            asin, pname_short, n_total, n_returns, n_comment,
            n_match, n_mis, n_unc, n_noc, n_refonly, mis_pct,
            top_reasons[0], top_reasons[1], top_reasons[2],
        ])

    MIS_PCT_COL, MIS_CNT_COL = 11, 7

    for row_idx in range(2, ws_sum.max_row + 1):
        mis_pct  = ws_sum.cell(row_idx, MIS_PCT_COL).value or 0
        row_fill = PatternFill('solid', fgColor='FFE0E0' if mis_pct >= 20 else
                               ('FFF3CC' if mis_pct >= 10 else 'F0F7F0'))
        for c in range(1, len(summary_cols) + 1):
            cell = ws_sum.cell(row_idx, c)
            cell.font   = body_font
            cell.border = border
            cell.alignment = center_align if c != 2 else left_align
            if c == MIS_PCT_COL:
                cell.number_format = '0.0"%"'
            cell.fill = (
                PatternFill('solid', fgColor='FF9999') if c == MIS_CNT_COL and mis_pct >= 20 else
                PatternFill('solid', fgColor=C_UNCLEAR) if c == MIS_CNT_COL and mis_pct >= 10 else
                PatternFill('solid', fgColor=C_REFUNDONLY) if c == 10 else
                row_fill
            )

    for i, w in enumerate([14, 52, 16, 14, 16, 8, 10, 8, 12, 15, 11, 24, 24, 24], 1):
        ws_sum.column_dimensions[get_column_letter(i)].width = w
    ws_sum.freeze_panes = 'A2'

    asin_order = (
        df.groupby('asin')['asin'].count()
        .sort_values(ascending=False).index.tolist()
    )

    DISPLAY_HEADERS = [
        'Return Date', 'ASIN', 'Product Name', 'Qty',
        'Return Reason (Amazon)', 'Customer Comment',
        'Analysis Status', 'True Reason', 'Comment Topic',
    ]

    for asin in asin_order:
        grp   = df[df['asin'] == asin].sort_values('return-date', ascending=False).reset_index(drop=True)
        ws    = wb.create_sheet(title=str(asin)[:31])
        pname = grp['product-name'].iloc[0]

        n_total   = len(grp)
        n_refonly = (grp['status'] == STATUS_REFUND_ONLY).sum()
        n_returns = n_total - n_refonly
        n_noc     = (grp['status'] == 'No Comment').sum()
        n_comment = n_returns - n_noc
        n_match   = (grp['status'] == 'Match').sum()
        n_mis     = (grp['status'] == 'Mismatch').sum()
        n_unc     = (grp['status'] == 'Unclear').sum()
        mis_pct   = round(n_mis / n_comment * 100, 1) if n_comment > 0 else 0

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
        tc = ws.cell(1, 1)
        tc.value     = f"{asin}  |  {pname[:120]}"
        tc.font      = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        tc.fill      = PatternFill('solid', fgColor=C_HEADER)
        tc.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[1].height = 22

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)
        sc = ws.cell(2, 1)
        sc.value = (
            f"Total Cases: {n_total}   |   Physical Returns: {n_returns}   |   "
            f"Refund w/o Return: {n_refonly}   |   With Comment: {n_comment}   |   "
            f"Match: {n_match}   |   Mismatch: {n_mis} ({mis_pct}%)   |   "
            f"Unclear: {n_unc}   |   No Comment: {n_noc}"
        )
        sc.font      = Font(name='Calibri', bold=True, size=9, color='1F3864')
        sc.fill      = PatternFill('solid', fgColor='DCE6F1')
        sc.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[2].height = 15

        analysis_end = write_analysis_block(ws, grp, start_row=3)

        header_row = analysis_end + 1
        data_start = header_row + 1

        for ci, hdr in enumerate(DISPLAY_HEADERS, 1):
            ws.cell(header_row, ci).value = hdr
        style_header_row(ws, header_row, len(DISPLAY_HEADERS))
        ws.freeze_panes = ws.cell(data_start, 1).coordinate

        alt = False
        for row_num, (_, row) in enumerate(grp.iterrows(), data_start):
            alt = not alt
            values = [
                row['return-date'],
                row['asin'],
                row['product-name'],
                row['quantity'],
                row['reason'],
                row['customer-comments'] if not pd.isna(row['customer-comments']) else '',
                row['status'],
                row['true_reason'],
                row['comment_topic'],
            ]
            status = row['status']
            for c_idx, val in enumerate(values, 1):
                cell            = ws.cell(row_num, c_idx)
                cell.value      = val
                cell.font       = body_font
                cell.border     = border
                cell.alignment  = wrap_align if c_idx == 6 else left_align
                if c_idx in (7, 8):
                    cell.fill = STATUS_FILL.get(status, PatternFill('solid', fgColor='FFFFFF'))
                elif c_idx == 5 and status == 'Mismatch':
                    cell.fill = PatternFill('solid', fgColor='FFB3B3')
                else:
                    cell.fill = PatternFill('solid', fgColor=C_ALT_ROW if alt else 'FFFFFF')
            ws.row_dimensions[row_num].height = 14

        for ci, w in enumerate([20, 12, 50, 5, 25, 55, 14, 24, 25, 10, 25], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        for r in range(data_start, ws.max_row + 1):
            ws.cell(r, 1).number_format = 'YYYY-MM-DD HH:MM'

    return wb


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--start", help="YYYY-MM-DD (required unless --local-returns)")
    parser.add_argument("--end", help="YYYY-MM-DD (required unless --local-returns)")
    parser.add_argument("--output", default="Returns_Analysis.xlsx", help="output .xlsx path")
    parser.add_argument("--save-raw", help="optional path to save the raw TSV report")
    parser.add_argument(
        "--transactions",
        help="Unified Transaction report CSV; adds refunds that have no physical "
             "return so totals match the Business Reports refund count",
    )
    parser.add_argument(
        "--local-returns",
        help="use this returns report file instead of calling SP-API (tab-separated TSV or XLSX)",
    )
    args = parser.parse_args()

    if args.local_returns:
        ext = os.path.splitext(args.local_returns)[1].lower()
        if ext in (".xlsx", ".xls"):
            raw = None
            df_local = pd.read_excel(args.local_returns, dtype=str)
        else:
            with open(args.local_returns, "rb") as f:
                data = f.read()
            try:
                raw = data.decode("utf-8")
            except UnicodeDecodeError:
                raw = data.decode("cp1252")
            df_local = None
    else:
        if not (args.start and args.end):
            parser.error("--start and --end are required unless --local-returns is used")
        cfg = get_config()
        raw = fetch_customer_returns(cfg, args.start, args.end)
        df_local = None

    if args.save_raw and raw is not None:
        with open(args.save_raw, "w", encoding="utf-8", newline="") as f:
            f.write(raw)
        print(f"Saved raw report to {args.save_raw}")

    if df_local is not None:
        df = normalize_returns_frame(df_local)
    else:
        df = parse_customer_returns(raw)
    if df.empty:
        sys.exit("No customer returns rows in report for this date range.")

    print(f"Parsed {len(df)} physical return rows across {df['asin'].nunique()} ASINs.")

    if args.transactions:
        extra = load_refund_only_rows(args.transactions, df)
        print(f"Adding {len(extra)} refunds with no matching physical return "
              f"(returnless refunds / never shipped back / not yet scanned).")
        df = pd.concat([df, extra], ignore_index=True)

    print("Building analysis workbook...")
    wb = build_workbook(df)
    wb.save(args.output)
    print(f"Saved: {args.output}")


if __name__ == "__main__":
    main()
