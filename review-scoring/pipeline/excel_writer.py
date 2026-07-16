"""Write the scoring workbook in the reference layout
(Positive / Negative / Usage / Improvement / Products).

Formatting mirrors the human-made reference workbook
(Styptic powder - 25.05.2026.xlsx): Arial throughout, a light-blue
header band, alternating light-gray banding per product within a
group, frozen header row, and matching column widths / font sizes.
"""
from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .models import Taxonomy, Group
from . import domain as _domain

FONT_NAME = "Arial"

HEADER_FILL = PatternFill("solid", fgColor="CFE2F3")
BAND_FILL = PatternFill("solid", fgColor="EFEFEF")

HEADER_FONT = Font(name=FONT_NAME, bold=True, size=11)
GROUP_FONT = Font(name=FONT_NAME, bold=True, size=11)
DETAIL_FONT = Font(name=FONT_NAME, size=10)
SUBTOTAL_FONT = Font(name=FONT_NAME, bold=True, size=11)
GRANDTOTAL_FONT = Font(name=FONT_NAME, bold=True, size=12)

HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")
TEXT_ALIGN = Alignment(wrap_text=True, vertical="center")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
CENTER_WRAP_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _set_widths(ws, widths: list[float]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _style_header(ws) -> None:
    for c in ws[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
    ws.freeze_panes = "A2"


def _sorted_groups(tax: Taxonomy, category: str) -> list[Group]:
    return sorted(tax.groups_for(category),
                  key=lambda g: (-g.total(tax), g.name))


def _write_two_level_sheet(ws, tax: Taxonomy, category: str,
                           headers: list[str], widths: list[float]) -> None:
    ws.append(headers)
    _style_header(ws)
    row = 2
    for group in _sorted_groups(tax, category):
        group_first = row
        per_product: dict[str, list] = defaultdict(list)
        for canon in group.canonicals(tax):
            for product, votes in canon.votes.items():
                if votes:
                    per_product[product].append((canon, votes))
        products = sorted(per_product,
                          key=lambda p: -sum(v for _, v in per_product[p]))
        for p_idx, product in enumerate(products):
            product_first = row
            band = BAND_FILL if p_idx % 2 == 1 else None
            entries = sorted(per_product[product], key=lambda cv: -cv[1])
            for canon, votes in entries:
                b = ws.cell(row=row, column=2, value=canon.text)
                b.font = DETAIL_FONT
                b.alignment = TEXT_ALIGN
                c = ws.cell(row=row, column=3, value=product)
                c.font = DETAIL_FONT
                c.alignment = CENTER_ALIGN
                d = ws.cell(row=row, column=4, value=votes)
                d.font = DETAIL_FONT
                d.alignment = CENTER_ALIGN
                if band:
                    for cell in (b, c, d):
                        cell.fill = band
                row += 1
            e = ws.cell(row=product_first, column=5,
                       value=f"=SUM(D{product_first}:D{row - 1})")
            e.font = SUBTOTAL_FONT
            e.alignment = CENTER_ALIGN
        if row > group_first:
            a = ws.cell(row=group_first, column=1, value=group.name)
            a.font = GROUP_FONT
            a.alignment = CENTER_WRAP_ALIGN
            f = ws.cell(row=group_first, column=6,
                       value=f"=SUM(E{group_first}:E{row - 1})")
            f.font = GRANDTOTAL_FONT
            f.alignment = CENTER_ALIGN
    _set_widths(ws, widths)


def _write_relation_sheet(ws, tax: Taxonomy, categories: list[str],
                          headers: list[str], widths: list[float],
                          bucket_labels: dict[str, str]) -> None:
    ws.append(headers)
    _style_header(ws)
    row = 2

    groups: list[Group] = []
    for cat in categories:
        groups += _sorted_groups(tax, cat)
    by_cat: dict[str, list[Group]] = defaultdict(list)
    for g in groups:
        cat = g.usage_category or bucket_labels.get(g.category, "Other")
        by_cat[cat].append(g)

    for cat, cat_groups in sorted(by_cat.items(),
                                  key=lambda kv: -sum(g.total(tax) for g in kv[1])):
        cat_cell = ws.cell(row=row, column=1, value=cat)
        cat_cell.font = GROUP_FONT
        cat_cell.fill = BAND_FILL
        cat_cell.alignment = CENTER_ALIGN
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 2
        for group in cat_groups:
            group_first = row
            for canon in sorted(group.canonicals(tax), key=lambda c: -c.total):
                canon_first = row
                for product, votes in sorted(canon.votes.items(),
                                             key=lambda kv: -kv[1]):
                    if not votes:
                        continue
                    b = ws.cell(row=row, column=2, value=canon.text)
                    b.font = DETAIL_FONT
                    b.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
                    c = ws.cell(row=row, column=3, value=product)
                    c.font = DETAIL_FONT
                    c.alignment = CENTER_ALIGN
                    d = ws.cell(row=row, column=4, value=votes)
                    d.font = SUBTOTAL_FONT
                    d.alignment = CENTER_ALIGN
                    row += 1
                e = ws.cell(row=canon_first, column=5,
                           value=f"=SUM(D{canon_first}:D{row - 1})")
                e.font = SUBTOTAL_FONT
                e.alignment = CENTER_ALIGN
                row += 1  # blank row between relations, as in the reference
            if row > group_first:
                a = ws.cell(row=group_first, column=1, value=group.name)
                a.font = GROUP_FONT
                a.alignment = CENTER_WRAP_ALIGN
                f = ws.cell(row=group_first, column=6,
                           value=f"=SUM(E{group_first}:E{row - 1})")
                f.font = GRANDTOTAL_FONT
                f.alignment = CENTER_ALIGN
    _set_widths(ws, widths)


def _write_wish_sheet(ws, tax: Taxonomy, category: str,
                      headers: list[str], widths: list[float]) -> None:
    ws.append(headers)
    _style_header(ws)
    row = 2
    for group in _sorted_groups(tax, category):
        group_first = row
        for canon in sorted(group.canonicals(tax), key=lambda c: -c.total):
            for product, votes in sorted(canon.votes.items(), key=lambda kv: -kv[1]):
                if not votes:
                    continue
                sample = "; ".join(canon.quotes.get(product, [])[:2])
                a = ws.cell(row=row, column=1, value=product)
                a.font = DETAIL_FONT
                a.alignment = Alignment(vertical="center")
                b = ws.cell(row=row, column=2,
                            value=group.name if canon.text == group.name else canon.text)
                b.font = DETAIL_FONT
                b.alignment = Alignment(wrap_text=True, vertical="center")
                c = ws.cell(row=row, column=3, value=sample)
                c.font = DETAIL_FONT
                c.alignment = TEXT_ALIGN
                d = ws.cell(row=row, column=4, value=votes)
                d.font = DETAIL_FONT
                d.alignment = CENTER_ALIGN
                row += 1
        if row > group_first:
            e = ws.cell(row=group_first, column=5,
                       value=f"=SUM(D{group_first}:D{row - 1})")
            e.font = GRANDTOTAL_FONT
            e.alignment = CENTER_ALIGN
            row += 1  # blank row between improvement themes
    _set_widths(ws, widths)


def write_workbook(tax: Taxonomy, products: dict[str, str],
                   out_path: Path) -> None:
    """products: short name -> link (for the Products sheet). The sheet
    layout is driven by the active domain profile (pipeline/domain.py);
    the default profile reproduces the reference workbook exactly."""
    from .grouping import reconcile_votes   # local: avoids import cycle
    reconcile_votes(tax)   # one review = one vote, on every write path
    dom = _domain.active()
    wb = Workbook()
    first = True
    for spec in dom.sheets:
        if first:
            ws = wb.active
            ws.title = spec.title
            first = False
        else:
            ws = wb.create_sheet(spec.title)
        if spec.layout == "two_level":
            _write_two_level_sheet(ws, tax, spec.categories[0],
                                   spec.headers, spec.widths)
        elif spec.layout == "relation":
            _write_relation_sheet(ws, tax, spec.categories, spec.headers,
                                  spec.widths, spec.bucket_labels)
        elif spec.layout == "wish":
            _write_wish_sheet(ws, tax, spec.categories[0],
                              spec.headers, spec.widths)
        else:
            raise ValueError(f"unknown sheet layout: {spec.layout}")

    ws = wb.create_sheet("Products")
    ws.append(["Product name", "Link"])
    _style_header(ws)
    for name, link in products.items():
        r = ws.max_row + 1
        ws.cell(row=r, column=1, value=name).font = DETAIL_FONT
        ws.cell(row=r, column=2, value=link).font = DETAIL_FONT
    _set_widths(ws, [19.75, 88.5])

    wb.save(out_path)
