"""Workbook writer: reconciles votes on write and emits correct SUM formulas."""
import tempfile
import unittest
from pathlib import Path
import openpyxl
from pipeline.excel_writer import write_workbook
from tests.helpers import tax_with, canon_by_text


class TestWorkbook(unittest.TestCase):
    def _write(self, tax):
        d = tempfile.mkdtemp()
        out = Path(d) / "wb.xlsx"
        write_workbook(tax, {"A": "", "B": ""}, out)
        return openpyxl.load_workbook(out)

    def test_write_reconciles_inflated_votes(self):
        tax = tax_with([("Easy to use", "positive", "easy to use",
                         {"A": ["A:1", "A:2"]})])
        canon_by_text(tax, "easy to use").votes["A"] = 9   # inflated
        wb = self._write(tax)
        ws = wb["Positive"]
        vals = [ws.cell(r, 4).value for r in range(2, ws.max_row + 1)
                if ws.cell(r, 4).value is not None]
        self.assertIn(2, vals)          # corrected to 2 distinct reviews
        self.assertNotIn(9, vals)

    def test_sum_formulas_present_and_scoped(self):
        tax = tax_with([
            ("Easy to use", "positive", "easy to use", {"A": ["A:1"], "B": ["B:1"]}),
        ])
        wb = self._write(tax)
        ws = wb["Positive"]
        formulas = [ws.cell(r, c).value for r in range(2, ws.max_row + 1)
                    for c in (5, 6)
                    if isinstance(ws.cell(r, c).value, str)
                    and ws.cell(r, c).value.startswith("=SUM")]
        self.assertTrue(any(f.startswith("=SUM(D") for f in formulas))  # per-product
        self.assertTrue(any(f.startswith("=SUM(E") for f in formulas))  # per-USP


if __name__ == "__main__":
    unittest.main()
