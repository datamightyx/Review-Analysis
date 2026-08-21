"""Excel workbook builder for Returns Analysis."""

import io
from typing import Any, List, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from services.constants import (
    KW,
    REASON_OK,
    TRUE_REASON_LABEL,
    TRIVIAL,
    TOPIC_DISPLAY,
    TOPIC_ROW_COLOR,
    STATUS_FILL_COLORS,
    C_HEADER,
    C_HEADER_FG,
    C_SUMMARY_H,
    C_ALT_ROW,
    C_REFUNDONLY,
    C_UNCLEAR,
    C_DISPO_BAD,
    C_DISPO_OK,
    REFUND_NO_RETURN,
    STATUS_REFUND_ONLY,
    ASIN_UNMAPPED,
    PRODUCT_UNKNOWN,
    DISPOSITION_DEFECTIVE,
    DISPOSITION_SELLABLE,
    DEFECT_CLAIM_REASONS,
    DISPO_CONFIRMED,
    DISPO_CONTRADICTED,
    DISPO_NA,
)


# Shared styles
body_font = Font(name='Calibri', size=9)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=False)
wrap_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center')

thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

STATUS_FILL = {
    status: PatternFill('solid', fgColor=color)
    for status, color in STATUS_FILL_COLORS.items()
}


def classify_comment(comment: Any) -> str:
    """Classify customer comment into topic category."""
    if pd.isna(comment):
        return 'NO_COMMENT'
    c = str(comment).strip().lower()
    if c in TRIVIAL or len(c) <= 2:
        return 'NO_COMMENT'
    for cat in [
        'SIZE_TOO_LARGE', 'SIZE_TOO_SMALL', 'DEFECTIVE', 'QUALITY_ISSUE',
        'NOT_AS_DESCRIBED', 'WRONG_ITEM', 'DELIVERY_ISSUE',
        'CHANGED_MIND', 'BETTER_PRICE', 'SIZE_ISSUE',
    ]:
        if any(k in c for k in KW[cat]):
            return cat
    return 'OTHER'


def get_status(row: pd.Series) -> str:
    """Determine analysis status for a row."""
    topic = row['comment_topic']
    if row['reason'] == REFUND_NO_RETURN:
        return STATUS_REFUND_ONLY
    if topic == 'NO_COMMENT':
        return 'No Comment'
    if topic == 'OTHER':
        return 'Unclear'
    expected = REASON_OK.get(str(row['reason']), [])
    return 'Match' if topic in expected else 'Mismatch'


def get_true_reason(row: pd.Series) -> str:
    """Get corrected true reason based on comment analysis."""
    if row['status'] == 'Mismatch':
        return TRUE_REASON_LABEL.get(row['comment_topic'], row['reason'])
    return row['reason']


def get_disposition_check(row: pd.Series) -> str:
    """Cross-check a claimed defect against what the warehouse physically found.

    detailed-disposition is Amazon's own grading of the returned unit, so unlike
    the comment-vs-reason mismatch it is an independent signal: a customer who
    claims DEFECTIVE on a unit graded SELLABLE is a different finding entirely.
    Only defect-family claims are checkable; everything else returns blank.
    """
    if row['reason'] == REFUND_NO_RETURN:
        return DISPO_NA
    if str(row['reason']).strip().upper() not in DEFECT_CLAIM_REASONS:
        return DISPO_NA
    dispo = row.get('detailed-disposition')
    if pd.isna(dispo):
        return DISPO_NA
    dispo = str(dispo).strip().upper()
    if dispo in DISPOSITION_DEFECTIVE:
        return DISPO_CONFIRMED
    if dispo in DISPOSITION_SELLABLE:
        return DISPO_CONTRADICTED
    return DISPO_NA


def _unique_sheet_title(wb: Workbook, base: Any) -> str:
    """Build a legal, unique worksheet title.

    Excel caps titles at 31 chars and forbids []:*?/\\, so unmapped-SKU
    sentinels and long SKUs can collide once truncated.
    """
    title = str(base)[:31].strip() or 'UNKNOWN'
    for ch in '[]:*?/\\':
        title = title.replace(ch, '-')
    if title not in wb.sheetnames:
        return title
    stem = title[:27]
    i = 2
    while f'{stem}~{i}' in wb.sheetnames:
        i += 1
    return f'{stem}~{i}'


def pct_bar(pct: float, width: int = 8) -> str:
    """Generate visual percentage bar."""
    filled = max(0, min(width, round(pct / 100 * width)))
    return '█' * filled + '░' * (width - filled)


def style_header_row(ws, row_num: int, col_count: int, bg_color: str = C_HEADER) -> None:
    """Style a header row."""
    fill = PatternFill('solid', fgColor=bg_color)
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill = fill
        cell.font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
        cell.alignment = center_align
        cell.border = border
    ws.row_dimensions[row_num].height = 18


def write_analysis_block(ws, grp: pd.DataFrame, start_row: int) -> int:
    """Write analysis block for a group (ASIN or overall)."""
    n_total = len(grp)
    n_noc = (grp['status'] == 'No Comment').sum()
    n_refonly = (grp['status'] == STATUS_REFUND_ONLY).sum()
    n_commented = n_total - n_noc - n_refonly
    n_mismatch = (grp['status'] == 'Mismatch').sum()

    topic_vc = (
        grp[grp['comment_topic'] != 'NO_COMMENT']['comment_topic']
        .value_counts()
    )
    true_vc = grp['true_reason'].value_counts()

    mis_grp = grp[grp['status'] == 'Mismatch']
    mismatch_rows = []
    if len(mis_grp) > 0:
        for stated, sub in mis_grp.groupby('reason'):
            mismatch_rows.append({
                'stated': stated,
                'count': len(sub),
                'top_true': sub['true_reason'].value_counts().index[0],
            })
        mismatch_rows.sort(key=lambda x: x['count'], reverse=True)

    n_data = min(max(len(topic_vc), len(true_vc), len(mismatch_rows), 1), 12)

    T1_COLOR = '375623'
    T2_COLOR = '1F3864'
    T3_COLOR = '7B2D00'

    def hfill(c):
        return PatternFill('solid', fgColor=c)

    def hdr_font(c='FFFFFF', sz=9, bold=True):
        return Font(name='Calibri', bold=bold, size=sz, color=c)

    def brd():
        s = Side(style='thin', color='CCCCCC')
        return Border(left=s, right=s, top=s, bottom=s)

    la = Alignment(horizontal='left', vertical='center')
    ca = Alignment(horizontal='center', vertical='center')

    def wc(row, col, val='', fill=None, font=None, align=None):
        cell = ws.cell(row, col)
        cell.value = val
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        if align:
            cell.alignment = align
        cell.border = brd()
        return cell

    r = start_row

    # Section headers
    for (c1, c2), color, label in [
        ((1, 3), T1_COLOR, 'COMMENT TOPICS  —  what customers actually write'),
        ((5, 7), T2_COLOR, 'TRUE RETURN REASONS  —  after mismatch correction'),
        ((9, 11), T3_COLOR, 'MISMATCH ANALYSIS  —  stated reason vs. comment'),
    ]:
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        wc(r, c1, label,
           fill=hfill(color),
           font=Font(name='Calibri', bold=True, size=10, color='FFFFFF'),
           align=Alignment(horizontal='left', vertical='center', indent=1))
    ws.row_dimensions[r].height = 18
    r += 1

    # Sub-headers
    sub_defs = [
        (1, 'Comment Topic', T1_COLOR),
        (2, 'Count', T1_COLOR),
        (3, '% / bar', T1_COLOR),
        (5, 'True Reason', T2_COLOR),
        (6, 'Count', T2_COLOR),
        (7, '% of Total', T2_COLOR),
        (9, 'Stated Reason (Amazon)', T3_COLOR),
        (10, '# Mismatches', T3_COLOR),
        (11, 'Corrected To', T3_COLOR),
    ]
    for col, label, color in sub_defs:
        wc(r, col, label, fill=hfill(color), font=hdr_font(sz=8), align=ca)
    ws.row_dimensions[r].height = 14
    r += 1

    t1_items = list(topic_vc.items())
    t2_items = list(true_vc.items())

    for i in range(n_data):
        alt = hfill('F7F7F7') if i % 2 else hfill('FFFFFF')
        df_fnt = Font(name='Calibri', size=9)
        sm_fnt = Font(name='Calibri', size=8)

        if i < len(t1_items):
            topic, cnt = t1_items[i]
            pct = cnt / n_commented * 100 if n_commented else 0
            rf = hfill(TOPIC_ROW_COLOR.get(topic, 'F2F2F2'))
            wc(r, 1, TOPIC_DISPLAY.get(topic, topic), fill=rf, font=df_fnt, align=la)
            wc(r, 2, cnt, fill=rf, font=df_fnt, align=ca)
            wc(r, 3, f'{pct:.1f}%  {pct_bar(pct)}', fill=rf, font=sm_fnt, align=la)
        else:
            for c in (1, 2, 3):
                wc(r, c, fill=hfill('FAFAFA'))

        if i < len(t2_items):
            reason, cnt = t2_items[i]
            pct = cnt / n_total * 100 if n_total else 0
            wc(r, 5, reason, fill=alt, font=df_fnt, align=la)
            wc(r, 6, cnt, fill=alt, font=df_fnt, align=ca)
            wc(r, 7, f'{pct:.1f}%  {pct_bar(pct)}', fill=alt, font=sm_fnt, align=la)
        else:
            for c in (5, 6, 7):
                wc(r, c, fill=hfill('FAFAFA'))

        if i < len(mismatch_rows):
            row_d = mismatch_rows[i]
            cnt = row_d['count']
            top_true = row_d['top_true']
            pct_m = cnt / n_mismatch * 100 if n_mismatch else 0
            wc(r, 9, row_d['stated'], fill=alt, font=df_fnt, align=la)
            wc(r, 10, f"{cnt}  ({pct_m:.0f}%)", fill=alt, font=df_fnt, align=ca)
            wc(r, 11, top_true,
               fill=hfill('FFD7D7'),
               font=Font(name='Calibri', size=9, bold=True, color='7B2D00'),
               align=la)
        else:
            for c in (9, 10, 11):
                wc(r, c, fill=hfill('FAFAFA'))

        ws.row_dimensions[r].height = 14
        r += 1

    # Totals row
    tf = hfill('ECECEC')
    tfnt = Font(name='Calibri', bold=True, size=9)
    sfnt = Font(name='Calibri', bold=True, size=8)

    n_with_topic = int(topic_vc.sum()) if len(topic_vc) else 0
    pct_topic = n_with_topic / n_commented * 100 if n_commented else 0
    wc(r, 1, 'TOTAL (with topic)', fill=tf, font=tfnt, align=la)
    wc(r, 2, n_with_topic, fill=tf, font=tfnt, align=ca)
    wc(r, 3, f'{pct_topic:.1f}% of {n_commented} commented', fill=tf, font=sfnt, align=la)

    wc(r, 5, 'TOTAL', fill=tf, font=tfnt, align=la)
    wc(r, 6, int(true_vc.sum()) if len(true_vc) else 0, fill=tf, font=tfnt, align=ca)
    wc(r, 7, '100%', fill=tf, font=sfnt, align=la)

    mis_rate = n_mismatch / n_commented * 100 if n_commented else 0
    wc(r, 9, 'TOTAL MISMATCHES', fill=tf, font=tfnt, align=la)
    wc(r, 10, n_mismatch, fill=tf, font=tfnt, align=ca)
    wc(r, 11, f'{mis_rate:.1f}% of commented returns', fill=tf, font=sfnt, align=la)

    ws.row_dimensions[r].height = 14
    r += 1
    return r


def build_workbook(df: pd.DataFrame) -> Workbook:
    """Build the analysis workbook from returns DataFrame."""
    df = df.copy()

    # Rows with no ASIN would be silently dropped by groupby, taking their units
    # with them. Park them under a visible sentinel instead.
    for col in ('detailed-disposition', 'status', 'refund-amount', 'refund-source'):
        if col not in df.columns:
            df[col] = 0.0 if col == 'refund-amount' else pd.NA
    df['asin'] = df['asin'].fillna(ASIN_UNMAPPED).replace('', ASIN_UNMAPPED)
    df['product-name'] = df['product-name'].fillna(PRODUCT_UNKNOWN)

    # Cases count rows; units count physical pieces. They are not the same number
    # and the unit figure is the one comparable to Business Reports unitsRefunded.
    df['units'] = pd.to_numeric(df['quantity'], errors='coerce').fillna(1)
    df['refund-amount'] = pd.to_numeric(df['refund-amount'], errors='coerce').fillna(0.0)

    # The report's own status column (Reimbursed / Units returned to inventory) is
    # kept before 'status' is reused for the analysis verdict.
    df['status_raw'] = df['status']

    df['comment_topic'] = df['customer-comments'].apply(classify_comment)
    df['status'] = df.apply(get_status, axis=1)
    df['true_reason'] = df.apply(get_true_reason, axis=1)
    df['dispo_check'] = df.apply(get_disposition_check, axis=1)

    wb = Workbook()

    # SUMMARY sheet
    summary_cols = [
        'ASIN', 'Product Name',
        'Total Cases (Returns + Refunds)', 'Physical Returns',
        'Units Total', 'Units Returned', 'Units Refund w/o Return',
        'Returns with Comment',
        'Match', 'Mismatch', 'Unclear', 'No Comment', 'Refund w/o Return',
        'Mismatch %',
        'Defect Claims', 'Defect Confirmed', 'Defect Contradicted',
        'Refund $ (no return)',
        'Top True Reason (Mismatch)',
        '2nd True Reason (Mismatch)',
        '3rd True Reason (Mismatch)',
    ]

    ws_sum = wb.active
    ws_sum.title = 'SUMMARY'
    ws_sum.append(summary_cols)
    style_header_row(ws_sum, 1, len(summary_cols), C_SUMMARY_H)

    for asin, grp in df.groupby('asin', sort=False):
        pname = str(grp['product-name'].iloc[0])
        pname_short = pname[:80] + '...' if len(pname) > 80 else pname
        n_total = len(grp)
        refonly_mask = grp['status'] == STATUS_REFUND_ONLY
        n_refonly = refonly_mask.sum()
        n_returns = n_total - n_refonly
        n_noc = (grp['status'] == 'No Comment').sum()
        n_comment = n_returns - n_noc
        n_match = (grp['status'] == 'Match').sum()
        n_mis = (grp['status'] == 'Mismatch').sum()
        n_unc = (grp['status'] == 'Unclear').sum()
        mis_pct = round(n_mis / n_comment * 100, 1) if n_comment > 0 else 0

        u_total = int(grp['units'].sum())
        u_refonly = int(grp.loc[refonly_mask, 'units'].sum())
        u_returns = u_total - u_refonly

        n_claim = (grp['dispo_check'] != DISPO_NA).sum()
        n_conf = (grp['dispo_check'] == DISPO_CONFIRMED).sum()
        n_contra = (grp['dispo_check'] == DISPO_CONTRADICTED).sum()
        refund_money = round(abs(grp.loc[refonly_mask, 'refund-amount'].sum()), 2)

        top_reasons = (
            grp[grp['status'] == 'Mismatch']['true_reason']
            .value_counts().head(3).index.tolist()
        )
        top_reasons += [''] * (3 - len(top_reasons))

        ws_sum.append([
            asin, pname_short, n_total, n_returns,
            u_total, u_returns, u_refonly,
            n_comment,
            n_match, n_mis, n_unc, n_noc, n_refonly, mis_pct,
            n_claim, n_conf, n_contra, refund_money,
            top_reasons[0], top_reasons[1], top_reasons[2],
        ])

    # Resolved by name so adding columns cannot silently shift the highlighting
    MIS_PCT_COL = summary_cols.index('Mismatch %') + 1
    MIS_CNT_COL = summary_cols.index('Mismatch') + 1
    REFONLY_COL = summary_cols.index('Refund w/o Return') + 1
    CONTRA_COL = summary_cols.index('Defect Contradicted') + 1
    MONEY_COL = summary_cols.index('Refund $ (no return)') + 1

    for row_idx in range(2, ws_sum.max_row + 1):
        mis_pct = ws_sum.cell(row_idx, MIS_PCT_COL).value or 0
        contra = ws_sum.cell(row_idx, CONTRA_COL).value or 0
        row_fill = PatternFill('solid', fgColor='FFE0E0' if mis_pct >= 20 else
                               ('FFF3CC' if mis_pct >= 10 else 'F0F7F0'))
        for c in range(1, len(summary_cols) + 1):
            cell = ws_sum.cell(row_idx, c)
            cell.font = body_font
            cell.border = border
            cell.alignment = center_align if c != 2 else left_align
            if c == MIS_PCT_COL:
                cell.number_format = '0.0"%"'
            if c == MONEY_COL:
                cell.number_format = '#,##0.00'
            cell.fill = (
                PatternFill('solid', fgColor='FF9999') if c == MIS_CNT_COL and mis_pct >= 20 else
                PatternFill('solid', fgColor=C_UNCLEAR) if c == MIS_CNT_COL and mis_pct >= 10 else
                PatternFill('solid', fgColor=C_DISPO_BAD) if c == CONTRA_COL and contra else
                PatternFill('solid', fgColor=C_REFUNDONLY) if c == REFONLY_COL else
                row_fill
            )

    col_widths = [
        14, 52, 16, 14, 11, 13, 18, 16, 8, 10, 8, 12, 15, 11,
        12, 15, 17, 17, 24, 24, 24,
    ]
    for i, w in enumerate(col_widths, 1):
        ws_sum.column_dimensions[get_column_letter(i)].width = w
    ws_sum.freeze_panes = 'A2'

    # Per-ASIN sheets
    asin_order = (
        df.groupby('asin')['asin'].count()
        .sort_values(ascending=False).index.tolist()
    )

    DISPLAY_HEADERS = [
        'Return Date', 'ASIN', 'Product Name', 'Qty',
        'Return Reason (Amazon)', 'Customer Comment',
        'Analysis Status', 'True Reason', 'Comment Topic',
        'Disposition', 'Return Status', 'Defect Check', 'Refund $',
    ]
    N_COLS = len(DISPLAY_HEADERS)

    for asin in asin_order:
        grp = df[df['asin'] == asin].sort_values('return-date', ascending=False).reset_index(drop=True)
        ws = wb.create_sheet(title=_unique_sheet_title(wb, asin))
        pname = str(grp['product-name'].iloc[0])

        n_total = len(grp)
        refonly_mask = grp['status'] == STATUS_REFUND_ONLY
        n_refonly = refonly_mask.sum()
        n_returns = n_total - n_refonly
        n_noc = (grp['status'] == 'No Comment').sum()
        n_comment = n_returns - n_noc
        n_match = (grp['status'] == 'Match').sum()
        n_mis = (grp['status'] == 'Mismatch').sum()
        n_unc = (grp['status'] == 'Unclear').sum()
        mis_pct = round(n_mis / n_comment * 100, 1) if n_comment > 0 else 0

        u_total = int(grp['units'].sum())
        u_refonly = int(grp.loc[refonly_mask, 'units'].sum())
        n_contra = (grp['dispo_check'] == DISPO_CONTRADICTED).sum()

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N_COLS)
        tc = ws.cell(1, 1)
        tc.value = f"{asin}  |  {pname[:120]}"
        tc.font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        tc.fill = PatternFill('solid', fgColor=C_HEADER)
        tc.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[1].height = 22

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=N_COLS)
        sc = ws.cell(2, 1)
        sc.value = (
            f"Total Cases: {n_total}   |   Physical Returns: {n_returns}   |   "
            f"Refund w/o Return: {n_refonly}   |   "
            f"Units: {u_total} (refund-only {u_refonly})   |   "
            f"With Comment: {n_comment}   |   "
            f"Match: {n_match}   |   Mismatch: {n_mis} ({mis_pct}%)   |   "
            f"Unclear: {n_unc}   |   No Comment: {n_noc}   |   "
            f"Defect contradicted by disposition: {n_contra}"
        )
        sc.font = Font(name='Calibri', bold=True, size=9, color='1F3864')
        sc.fill = PatternFill('solid', fgColor='DCE6F1')
        sc.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[2].height = 15

        analysis_end = write_analysis_block(ws, grp, start_row=3)

        header_row = analysis_end + 1
        data_start = header_row + 1

        for ci, hdr in enumerate(DISPLAY_HEADERS, 1):
            ws.cell(header_row, ci).value = hdr
        style_header_row(ws, header_row, len(DISPLAY_HEADERS))
        ws.freeze_panes = ws.cell(data_start, 1).coordinate

        alt = False
        for row_num, (_, row) in enumerate(grp.iterrows(), data_start):
            alt = not alt
            dispo_check = row['dispo_check']
            values = [
                row['return-date'],
                row['asin'],
                row['product-name'],
                row['quantity'],
                row['reason'],
                row['customer-comments'] if not pd.isna(row['customer-comments']) else '',
                row['status'],
                row['true_reason'],
                row['comment_topic'],
                '' if pd.isna(row['detailed-disposition']) else row['detailed-disposition'],
                '' if pd.isna(row['status_raw']) else row['status_raw'],
                dispo_check,
                round(abs(row['refund-amount']), 2) or '',
            ]
            status = row['status']
            for c_idx, val in enumerate(values, 1):
                cell = ws.cell(row_num, c_idx)
                cell.value = val
                cell.font = body_font
                cell.border = border
                cell.alignment = wrap_align if c_idx == 6 else left_align
                if c_idx in (7, 8):
                    cell.fill = STATUS_FILL.get(status, PatternFill('solid', fgColor='FFFFFF'))
                elif c_idx == 5 and status == 'Mismatch':
                    cell.fill = PatternFill('solid', fgColor='FFB3B3')
                elif c_idx == 12 and dispo_check == DISPO_CONTRADICTED:
                    cell.fill = PatternFill('solid', fgColor=C_DISPO_BAD)
                elif c_idx == 12 and dispo_check == DISPO_CONFIRMED:
                    cell.fill = PatternFill('solid', fgColor=C_DISPO_OK)
                else:
                    cell.fill = PatternFill('solid', fgColor=C_ALT_ROW if alt else 'FFFFFF')
            ws.row_dimensions[row_num].height = 14

        detail_widths = [20, 12, 50, 5, 25, 55, 14, 24, 25, 20, 22, 26, 11]
        for ci, w in enumerate(detail_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        for r in range(data_start, ws.max_row + 1):
            ws.cell(r, 1).number_format = 'YYYY-MM-DD HH:MM'
            ws.cell(r, 13).number_format = '#,##0.00'

    return wb


def save_workbook(wb: Workbook, filepath: str) -> None:
    """Save workbook to file."""
    wb.save(filepath)


def workbook_to_bytes(wb: Workbook) -> bytes:
    """Convert workbook to bytes for download."""
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()