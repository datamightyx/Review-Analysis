"""Improvement groups are named by the short `gist` wish label, not by the
raw customer sentence a judge echoes back — and that label is actually
visible on the Improvement sheet.

Without this the wish sheet degenerates into one group per quote (measured
on the Hamster Sand workbook 2026-08-17: 12 groups, 8 named by a full
sentence, two pairs of them the same wish), and the consolidation pass has no
theme label left to match on.
"""
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from pipeline.grouping import _group_name_for
from pipeline.excel_writer import write_workbook
from helpers import tax_with

LONG_QUOTE = ("I just with it came with a larger quantity for the price "
              "they are asking")


def bucket(gist="", text=LONG_QUOTE):
    return {"gist": gist, "text": text}


class TestGistGroupName(unittest.TestCase):
    def test_long_proposed_name_loses_to_the_gist(self):
        self.assertEqual(
            _group_name_for("improvement", bucket("Bigger amount"), LONG_QUOTE),
            "Bigger amount")

    def test_short_proposed_name_wins(self):
        """The grouping judge sees the whole taxonomy, so a short name it
        proposes may be an EXISTING theme label — prefer it over the gist."""
        self.assertEqual(
            _group_name_for("improvement", bucket("Bigger amount"),
                            "Bigger bag"),
            "Bigger bag")

    def test_gist_used_when_the_judge_proposed_nothing(self):
        self.assertEqual(
            _group_name_for("improvement", bucket("Silica free sand"), ""),
            "Silica free sand")

    def test_phrase_text_when_there_is_no_gist(self):
        self.assertEqual(_group_name_for("improvement", bucket(""), ""),
                         LONG_QUOTE)

    def test_other_categories_are_unaffected(self):
        """positive/negative are not named from a gist — a proposed name
        always wins there, exactly as before."""
        self.assertEqual(
            _group_name_for("positive", bucket("Bigger amount"), LONG_QUOTE),
            LONG_QUOTE)
        self.assertEqual(_group_name_for("positive", bucket("x"), ""),
                         LONG_QUOTE)

    def test_theme_label_reaches_the_improvement_sheet(self):
        tax = tax_with([
            ("Bigger amount", "improvement", "I wish it was bigger",
             {"Sukh": ["r1", "r2"]}),
            ("Bigger amount", "improvement", "wish the bag was larger",
             {"Sukh": ["r3"]}),
        ])
        with tempfile.TemporaryDirectory() as d:
            out = Path(d) / "wb.xlsx"
            write_workbook(tax, {"Sukh": ""}, out)
            ws = load_workbook(out)["Improvement"]
            col_b = [ws.cell(row=r, column=2).value
                     for r in range(2, ws.max_row + 1)]
        # first row of the group carries the theme, later rows their own
        # wording — and the theme is not silently replaced by a quote
        self.assertEqual(col_b[0], "Bigger amount")
        self.assertIn("wish the bag was larger", col_b)


if __name__ == "__main__":
    unittest.main()
