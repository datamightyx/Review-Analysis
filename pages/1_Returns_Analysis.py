import io
import re

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Returns Analysis", page_icon="📦", layout="wide")
st.title("📦 Returns Analysis")
st.markdown("---")

# Звіт по поверненнях містить лише юніти, які фізично приїхали на FC і були
# відскановані. Refunds (Business Reports / Unified Transaction) рахують
# повернені гроші й додатково включають returnless refunds, товар, який покупець
# не відправив назад, та повернення ще в дорозі. Тому refunds >= returns.
# Другий (необов'язковий) файл — Unified Transaction — дозволяє додати refund-и
# без фізичного повернення, щоб підсумки збігалися з цифрою refunds.
REFUND_NO_RETURN   = 'REFUND_NO_RETURN'
STATUS_REFUND_ONLY = 'Refund - No Return'
RECON_KEY_COLS     = ['order-id', 'sku']


def load_refund_only_rows(tx_file, returns_df):
    """Refund-події з Unified Transaction CSV, яким немає пари в поверненнях."""
    raw = tx_file.getvalue() if hasattr(tx_file, 'getvalue') else tx_file.read()
    for enc in ('utf-8', 'utf-8-sig', 'cp1252', 'latin-1'):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw.decode('utf-8', errors='replace')

    # зріз по сирому тексту, не splitlines(): рядки даних містять поля в лапках
    # з переносами всередині
    m = re.search(r'^"?date/time"?\s*,', text, re.IGNORECASE | re.MULTILINE)
    if m is None:
        raise ValueError("не знайдено рядок заголовка 'date/time'")

    tx = pd.read_csv(io.StringIO(text[m.start():]), dtype=str)
    tx.columns = [c.strip() for c in tx.columns]
    for col in ('type', 'order id', 'sku', 'quantity', 'date/time'):
        if col not in tx.columns:
            raise ValueError(f"у файлі транзакцій немає колонки '{col}'")

    tx['quantity'] = pd.to_numeric(tx['quantity'], errors='coerce').fillna(0)
    ref = tx[tx['type'].astype(str).str.strip().str.lower() == 'refund'].copy()
    ref = ref[ref['quantity'] > 0]  # quantity 0 == рядок лише з комісією/податком
    ref['dt'] = pd.to_datetime(
        ref['date/time'].str.replace(r'\s+P[DS]T$', '', regex=True),
        format='%b %d, %Y %I:%M:%S %p',
        errors='coerce',
    )

    sku_asin = (
        returns_df.dropna(subset=['sku', 'asin'])
        .groupby('sku')['asin'].agg(lambda s: s.value_counts().index[0]).to_dict()
    )
    sku_name = (
        returns_df.dropna(subset=['sku', 'product-name'])
        .groupby('sku')['product-name'].first().to_dict()
    )

    ret_keys = set(zip(returns_df['order-id'], returns_df['sku']))
    ref = ref[[(o, s) not in ret_keys for o, s in zip(ref['order id'], ref['sku'])]]

    skipped = sorted(ref[~ref['sku'].isin(sku_asin)]['sku'].dropna().unique().tolist())
    ref = ref[ref['sku'].isin(sku_asin)]

    extra = pd.DataFrame({
        'return-date':       ref['dt'].values,
        'order-id':          ref['order id'].values,
        'sku':               ref['sku'].values,
        'asin':              ref['sku'].map(sku_asin).values,
        'product-name':      ref['sku'].map(sku_name).values,
        'quantity':          ref['quantity'].values,
        'reason':            REFUND_NO_RETURN,
        'customer-comments': pd.NA,
    })
    return extra, skipped

# ── 0. Template download ───────────────────────────────────────────────────────
def build_template() -> bytes:
    wb_t = Workbook()
    ws_t = wb_t.active
    ws_t.title = 'ALL'
    headers = [
        'return-date', 'order-id', 'sku', 'asin', 'product-name', 'quantity',
        'reason', 'customer-comments',
    ]
    ws_t.append(headers)
    hdr_fill = PatternFill('solid', fgColor='1F3864')
    hdr_font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    thin = Side(style='thin', color='CCCCCC')
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws_t.cell(1, col_idx)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = brd
    col_widths = [18, 22, 22, 14, 50, 8, 30, 60]
    for i, w in enumerate(col_widths, 1):
        ws_t.column_dimensions[get_column_letter(i)].width = w
    ws_t.row_dimensions[1].height = 18
    # example row
    ws_t.append([
        '2024-01-15', '111-0000000-0000000', 'SKU-FBA',
        'B0XXXXXXXXX', 'Example Product Name', 1,
        'DEFECTIVE', 'The item stopped working after two days.',
    ])
    buf = io.BytesIO()
    wb_t.save(buf)
    buf.seek(0)
    return buf.read()

with st.expander("Шаблон файлу (.xlsx)", expanded=False):
    st.markdown(
        "Файл повинен містити аркуш **`ALL`** з колонками нижче. "
        "Поля **`return-date`** та **`quantity`** — необов'язкові. "
        "**`order-id`** і **`sku`** потрібні лише для звірки з refunds "
        "(другий файл — Unified Transaction)."
    )
    st.table(pd.DataFrame({
        'Колонка': ['return-date', 'order-id', 'sku', 'asin', 'product-name',
                    'quantity', 'reason', 'customer-comments'],
        'Обов\'язкова': ['Ні', 'Для звірки', 'Для звірки', 'Так', 'Так', 'Ні', 'Так', 'Так'],
        'Приклад': ['2024-01-15', '111-0000000-0000000', 'SKU-FBA', 'B0XXXXXXXXX',
                    'Example Product', '1', 'DEFECTIVE', 'Stopped working'],
    }))
    st.download_button(
        label="Завантажити шаблон",
        data=build_template(),
        file_name="returns_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

st.markdown("---")

# ── 1. File upload ────────────────────────────────────────────────────────────
uploaded = st.file_uploader("Завантажте файл з поверненнями (.xlsx)", type=["xlsx"])

uploaded_tx = st.file_uploader(
    "Необов'язково: Unified Transaction report (.csv) — щоб додати refund-и без фізичного повернення",
    type=["csv"],
)

if not uploaded:
    st.info("Завантажте файл, щоб продовжити.")
    st.stop()

# ── 2. Read data & show ASIN selector ─────────────────────────────────────────
try:
    df_all = pd.read_excel(uploaded, sheet_name='ALL')
except Exception as e:
    st.error(f"Не вдалося прочитати файл: {e}")
    st.stop()

# Normalize column names: strip whitespace and lowercase
df_all.columns = df_all.columns.str.strip().str.lower()

REQUIRED_COLS = ['asin', 'product-name', 'reason', 'customer-comments']
missing = [c for c in REQUIRED_COLS if c not in df_all.columns]
if missing:
    st.error(
        f"У файлі відсутні обов'язкові колонки: **{missing}**\n\n"
        f"Знайдені колонки: `{list(df_all.columns)}`"
    )
    st.stop()

if 'return-date' not in df_all.columns:
    df_all['return-date'] = pd.NaT
else:
    df_all['return-date'] = pd.to_datetime(df_all['return-date'], errors='coerce', utc=True).dt.tz_localize(None)

if 'quantity' not in df_all.columns:
    df_all['quantity'] = None

recon_ready = all(c in df_all.columns for c in RECON_KEY_COLS)
if uploaded_tx and not recon_ready:
    st.warning(
        f"Звірку з refunds пропущено: у файлі повернень немає колонок "
        f"**{[c for c in RECON_KEY_COLS if c not in df_all.columns]}**. "
        "Вони потрібні для join з транзакціями."
    )

all_asins = sorted(df_all['asin'].dropna().unique().tolist())

st.subheader("Оберіть ASIN для аналізу")
col1, col2 = st.columns([1, 4])
with col1:
    select_all = st.checkbox("Вибрати всі", value=False)

if select_all:
    selected_asins = all_asins
else:
    selected_asins = st.multiselect(
        "ASIN коди:",
        options=all_asins,
        default=[],
        placeholder="Оберіть один або кілька ASIN...",
    )

if not selected_asins:
    st.warning("Оберіть хоча б один ASIN.")
    st.stop()

st.markdown(f"Обрано: **{len(selected_asins)}** ASIN(s)")

# ── 3. Analyze button ─────────────────────────────────────────────────────────
if not st.button("▶ Аналізувати", type="primary"):
    st.stop()

with st.spinner("Аналіз..."):
    df = df_all[df_all['asin'].isin(selected_asins)].copy()

    n_refund_added, skipped_skus = 0, []
    if uploaded_tx and recon_ready:
        try:
            extra, skipped_skus = load_refund_only_rows(uploaded_tx, df_all)
            extra = extra[extra['asin'].isin(selected_asins)]
            n_refund_added = len(extra)
            df = pd.concat([df, extra], ignore_index=True)
        except Exception as e:
            st.error(f"Не вдалося прочитати файл транзакцій: {e}")
            st.stop()

    # ── Keyword sets ──────────────────────────────────────────────────────────
    KW = {
        'SIZE_TOO_LARGE': [
            'too large','too big','too wide','too long','too tall','runs large',
            'incorrect size','wrong size','size is wrong','too bulky','size too big',
            'bigger than','larger than','size larger','bit too big',
        ],
        'SIZE_TOO_SMALL': [
            'too small','too tight','too narrow','too short','runs small','size too small',
            'smaller than','too tiny','not big enough','too little','size smaller',
        ],
        'DEFECTIVE': [
            'defective','broken',"doesn't work",'does not work','not work','broke',
            'malfunction','stopped working','falling off','falls off','cracked',
            'leaking',' leak ','torn','ripped','damaged','doesnt function',
            "doesn't function",'no longer work',
        ],
        'QUALITY_ISSUE': [
            'poor quality','bad quality','too thin','very thin','flimsy','not durable',
            'low quality','inferior','terrible quality','bad material',
            'quality is bad','not good quality','not the same quality','cheap quality',
            'quality is poor','material is thin',
        ],
        'NOT_AS_DESCRIBED': [
            'not as described','not as expected','not what i expected',
            'misleading','misrepresented','inaccurate description',
        ],
        'WRONG_ITEM': [
            'wrong item','wrong product','wrong color','ordered wrong',
        ],
        'DELIVERY_ISSUE': [
            'never arrived','did not arrive','not delivered','never received',
            'late delivery','delayed','shipping was','weeks delayed','not arrive',
            'took too long','not received',
        ],
        'CHANGED_MIND': [
            'changed mind',"don't need",'do not need','no longer need',
            'no longer want','changed my mind',"don't want",'dont need','dont want',
        ],
        'BETTER_PRICE': [
            'found cheaper','better price','found better price','cheaper elsewhere',
            'cheaper on','lower price',
        ],
        'SIZE_ISSUE': [
            ' size ',' fit ',' fits ','fitting','incorrect size','wrong size',
            "doesn't fit","does not fit","wont fit","won't fit",
        ],
    }

    REASON_OK = {
        'APPAREL_TOO_LARGE':         ['SIZE_TOO_LARGE','SIZE_ISSUE'],
        'APPAREL_TOO_SMALL':         ['SIZE_TOO_SMALL','SIZE_ISSUE'],
        'POOR_FIT':                  ['SIZE_TOO_LARGE','SIZE_TOO_SMALL','SIZE_ISSUE'],
        'DEFECTIVE':                 ['DEFECTIVE'],
        'NOT_AS_DESCRIBED':          ['NOT_AS_DESCRIBED','QUALITY_ISSUE','SIZE_TOO_LARGE','SIZE_TOO_SMALL'],
        'QUALITY_UNACCEPTABLE':      ['QUALITY_ISSUE','DEFECTIVE'],
        'ORDERED_WRONG_ITEM':        ['WRONG_ITEM'],
        'UNWANTED_ITEM':             ['CHANGED_MIND','BETTER_PRICE'],
        'FOUND_BETTER_PRICE':        ['BETTER_PRICE','CHANGED_MIND'],
        'MISSING_PARTS':             ['DEFECTIVE'],
        'DAMAGED_BY_FC':             ['DEFECTIVE'],
        'DAMAGED_BY_CARRIER':        ['DEFECTIVE'],
        'NEVER_ARRIVED':             ['DELIVERY_ISSUE'],
        'MISSED_ESTIMATED_DELIVERY': ['DELIVERY_ISSUE'],
        'UNDELIVERABLE_UNKNOWN':     ['DELIVERY_ISSUE'],
        'UNDELIVERABLE_REFUSED':     ['DELIVERY_ISSUE'],
        'NO_REASON_GIVEN':           [],
        'SWITCHEROO':                ['WRONG_ITEM'],
        'NOT_COMPATIBLE':            ['NOT_AS_DESCRIBED','SIZE_ISSUE'],
        'PART_NOT_COMPATIBLE':       ['NOT_AS_DESCRIBED','SIZE_ISSUE'],
        'MISORDERED':                ['WRONG_ITEM'],
        'EXTRA_ITEM':                [],
        'EXCESSIVE_INSTALLATION':    [],
        'UNAUTHORIZED_PURCHASE':     [],
    }

    TRUE_REASON_LABEL = {
        'SIZE_TOO_LARGE':   'SIZE - Too Large',
        'SIZE_TOO_SMALL':   'SIZE - Too Small',
        'SIZE_ISSUE':       'SIZE - Fit Issue',
        'DEFECTIVE':        'DEFECTIVE',
        'QUALITY_ISSUE':    'QUALITY ISSUE',
        'NOT_AS_DESCRIBED': 'NOT AS DESCRIBED',
        'WRONG_ITEM':       'WRONG ITEM',
        'DELIVERY_ISSUE':   'DELIVERY ISSUE',
        'CHANGED_MIND':     'CHANGED MIND',
        'BETTER_PRICE':     'FOUND BETTER PRICE',
    }

    TRIVIAL = {'', 'na', 'n/a', 'no', 'yes', 'z', 'return', 'ok', 'none', '.', '-', 'n'}

    def classify_comment(comment):
        if pd.isna(comment):
            return 'NO_COMMENT'
        c = str(comment).strip().lower()
        if c in TRIVIAL or len(c) <= 2:
            return 'NO_COMMENT'
        for cat in [
            'SIZE_TOO_LARGE','SIZE_TOO_SMALL','DEFECTIVE','QUALITY_ISSUE',
            'NOT_AS_DESCRIBED','WRONG_ITEM','DELIVERY_ISSUE',
            'CHANGED_MIND','BETTER_PRICE','SIZE_ISSUE',
        ]:
            if any(k in c for k in KW[cat]):
                return cat
        return 'OTHER'

    def get_status(row):
        topic = row['comment_topic']
        if row['reason'] == REFUND_NO_RETURN:
            return STATUS_REFUND_ONLY
        if topic == 'NO_COMMENT':
            return 'No Comment'
        if topic == 'OTHER':
            return 'Unclear'
        expected = REASON_OK.get(str(row['reason']), [])
        return 'Match' if topic in expected else 'Mismatch'

    def get_true_reason(row):
        if row['status'] == 'Mismatch':
            return TRUE_REASON_LABEL.get(row['comment_topic'], row['reason'])
        return row['reason']

    df['comment_topic'] = df['customer-comments'].apply(classify_comment)
    df['status']        = df.apply(get_status, axis=1)
    df['true_reason']   = df.apply(get_true_reason, axis=1)

    # ── Build Excel ───────────────────────────────────────────────────────────
    C_HEADER    = '1F3864'
    C_HEADER_FG = 'FFFFFF'
    C_MISMATCH  = 'FFD7D7'
    C_MATCH     = 'D7F0D7'
    C_UNCLEAR   = 'FFF3CC'
    C_NOCOMMENT = 'F2F2F2'
    C_SUMMARY_H = '2E75B6'
    C_ALT_ROW   = 'EEF3FA'

    C_REFUNDONLY = 'E4DFEC'

    STATUS_FILL = {
        'Mismatch':         PatternFill('solid', fgColor=C_MISMATCH),
        'Match':            PatternFill('solid', fgColor=C_MATCH),
        'Unclear':          PatternFill('solid', fgColor=C_UNCLEAR),
        'No Comment':       PatternFill('solid', fgColor=C_NOCOMMENT),
        STATUS_REFUND_ONLY: PatternFill('solid', fgColor=C_REFUNDONLY),
    }

    header_font  = Font(name='Calibri', bold=True, color=C_HEADER_FG, size=10)
    body_font    = Font(name='Calibri', size=9)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=False)
    wrap_align   = Alignment(horizontal='left',   vertical='top',    wrap_text=True)
    left_align   = Alignment(horizontal='left',   vertical='center')

    thin   = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    TOPIC_DISPLAY = {
        'SIZE_TOO_LARGE':   'SIZE — Too Large',
        'SIZE_TOO_SMALL':   'SIZE — Too Small',
        'SIZE_ISSUE':       'SIZE — Fit Issue',
        'DEFECTIVE':        'Defective / Not Working',
        'QUALITY_ISSUE':    'Quality Issue',
        'NOT_AS_DESCRIBED': 'Not As Described',
        'WRONG_ITEM':       'Wrong Item Sent',
        'DELIVERY_ISSUE':   'Delivery Issue',
        'CHANGED_MIND':     'Changed Mind',
        'BETTER_PRICE':     'Found Better Price',
        'OTHER':            'Other / Unclear',
    }

    TOPIC_ROW_COLOR = {
        'SIZE_TOO_LARGE':   'FDE9D9',
        'SIZE_TOO_SMALL':   'FDE9D9',
        'SIZE_ISSUE':       'FDE9D9',
        'DEFECTIVE':        'FFD7D7',
        'QUALITY_ISSUE':    'EAD1DC',
        'NOT_AS_DESCRIBED': 'D9E1F2',
        'WRONG_ITEM':       'FCE4D6',
        'DELIVERY_ISSUE':   'DDEBF7',
        'CHANGED_MIND':     'E2EFDA',
        'BETTER_PRICE':     'D9F2F2',
        'OTHER':            'F2F2F2',
    }

    def pct_bar(pct, width=8):
        filled = max(0, min(width, round(pct / 100 * width)))
        return '█' * filled + '░' * (width - filled)

    def style_header_row(ws, row_num, col_count, bg_color=C_HEADER):
        fill = PatternFill('solid', fgColor=bg_color)
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill = fill
            cell.font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
            cell.alignment = center_align
            cell.border = border
        ws.row_dimensions[row_num].height = 18

    def write_analysis_block(ws, grp, start_row):
        n_total     = len(grp)
        n_noc       = (grp['status'] == 'No Comment').sum()
        n_refonly   = (grp['status'] == STATUS_REFUND_ONLY).sum()
        n_commented = n_total - n_noc - n_refonly
        n_mismatch  = (grp['status'] == 'Mismatch').sum()

        topic_vc = (
            grp[grp['comment_topic'] != 'NO_COMMENT']['comment_topic']
            .value_counts()
        )
        true_vc = grp['true_reason'].value_counts()

        mis_grp = grp[grp['status'] == 'Mismatch']
        mismatch_rows = []
        if len(mis_grp) > 0:
            for stated, sub in mis_grp.groupby('reason'):
                mismatch_rows.append({
                    'stated':   stated,
                    'count':    len(sub),
                    'top_true': sub['true_reason'].value_counts().index[0],
                })
            mismatch_rows.sort(key=lambda x: x['count'], reverse=True)

        n_data = min(max(len(topic_vc), len(true_vc), len(mismatch_rows), 1), 12)

        T1_COLOR = '375623'
        T2_COLOR = '1F3864'
        T3_COLOR = '7B2D00'

        def hfill(c):
            return PatternFill('solid', fgColor=c)

        def hdr_font(c='FFFFFF', sz=9, bold=True):
            return Font(name='Calibri', bold=bold, size=sz, color=c)

        def brd():
            s = Side(style='thin', color='CCCCCC')
            return Border(left=s, right=s, top=s, bottom=s)

        la = Alignment(horizontal='left',   vertical='center')
        ca = Alignment(horizontal='center', vertical='center')

        def wc(row, col, val='', fill=None, font=None, align=None):
            cell = ws.cell(row, col)
            cell.value = val
            if fill:  cell.fill      = fill
            if font:  cell.font      = font
            if align: cell.alignment = align
            cell.border = brd()
            return cell

        r = start_row

        for (c1, c2), color, label in [
            ((1, 3), T1_COLOR, 'COMMENT TOPICS  —  what customers actually write'),
            ((5, 7), T2_COLOR, 'TRUE RETURN REASONS  —  after mismatch correction'),
            ((9,11), T3_COLOR, 'MISMATCH ANALYSIS  —  stated reason vs. comment'),
        ]:
            ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
            wc(r, c1, label,
               fill=hfill(color),
               font=Font(name='Calibri', bold=True, size=10, color='FFFFFF'),
               align=Alignment(horizontal='left', vertical='center', indent=1))
        ws.row_dimensions[r].height = 18
        r += 1

        sub_defs = [
            (1,  'Comment Topic',           T1_COLOR),
            (2,  'Count',                   T1_COLOR),
            (3,  '% / bar',                 T1_COLOR),
            (5,  'True Reason',             T2_COLOR),
            (6,  'Count',                   T2_COLOR),
            (7,  '% of Total',              T2_COLOR),
            (9,  'Stated Reason (Amazon)',  T3_COLOR),
            (10, '# Mismatches',            T3_COLOR),
            (11, 'Corrected To',            T3_COLOR),
        ]
        for col, label, color in sub_defs:
            wc(r, col, label, fill=hfill(color), font=hdr_font(sz=8), align=ca)
        ws.row_dimensions[r].height = 14
        r += 1

        t1_items = list(topic_vc.items())
        t2_items = list(true_vc.items())

        for i in range(n_data):
            alt    = hfill('F7F7F7') if i % 2 else hfill('FFFFFF')
            df_fnt = Font(name='Calibri', size=9)
            sm_fnt = Font(name='Calibri', size=8)

            if i < len(t1_items):
                topic, cnt = t1_items[i]
                pct = cnt / n_commented * 100 if n_commented else 0
                rf  = hfill(TOPIC_ROW_COLOR.get(topic, 'F2F2F2'))
                wc(r, 1, TOPIC_DISPLAY.get(topic, topic), fill=rf, font=df_fnt, align=la)
                wc(r, 2, cnt,                              fill=rf, font=df_fnt, align=ca)
                wc(r, 3, f'{pct:.1f}%  {pct_bar(pct)}',   fill=rf, font=sm_fnt, align=la)
            else:
                for c in (1, 2, 3):
                    wc(r, c, fill=hfill('FAFAFA'))

            if i < len(t2_items):
                reason, cnt = t2_items[i]
                pct = cnt / n_total * 100 if n_total else 0
                wc(r, 5, reason,                         fill=alt, font=df_fnt, align=la)
                wc(r, 6, cnt,                            fill=alt, font=df_fnt, align=ca)
                wc(r, 7, f'{pct:.1f}%  {pct_bar(pct)}', fill=alt, font=sm_fnt, align=la)
            else:
                for c in (5, 6, 7):
                    wc(r, c, fill=hfill('FAFAFA'))

            if i < len(mismatch_rows):
                row_d    = mismatch_rows[i]
                cnt      = row_d['count']
                top_true = row_d['top_true']
                pct_m    = cnt / n_mismatch * 100 if n_mismatch else 0
                wc(r,  9, row_d['stated'],               fill=alt, font=df_fnt, align=la)
                wc(r, 10, f"{cnt}  ({pct_m:.0f}%)",      fill=alt, font=df_fnt, align=ca)
                wc(r, 11, top_true,
                   fill=hfill('FFD7D7'),
                   font=Font(name='Calibri', size=9, bold=True, color='7B2D00'),
                   align=la)
            else:
                for c in (9, 10, 11):
                    wc(r, c, fill=hfill('FAFAFA'))

            ws.row_dimensions[r].height = 14
            r += 1

        tf   = hfill('ECECEC')
        tfnt = Font(name='Calibri', bold=True, size=9)
        sfnt = Font(name='Calibri', bold=True, size=8)

        n_with_topic = int(topic_vc.sum()) if len(topic_vc) else 0
        pct_topic    = n_with_topic / n_commented * 100 if n_commented else 0
        wc(r, 1, 'TOTAL (with topic)',                       fill=tf, font=tfnt, align=la)
        wc(r, 2, n_with_topic,                               fill=tf, font=tfnt, align=ca)
        wc(r, 3, f'{pct_topic:.1f}% of {n_commented} commented', fill=tf, font=sfnt, align=la)

        wc(r, 5, 'TOTAL',                                    fill=tf, font=tfnt, align=la)
        wc(r, 6, int(true_vc.sum()) if len(true_vc) else 0,  fill=tf, font=tfnt, align=ca)
        wc(r, 7, '100%',                                     fill=tf, font=sfnt, align=la)

        mis_rate = n_mismatch / n_commented * 100 if n_commented else 0
        wc(r,  9, 'TOTAL MISMATCHES',                        fill=tf, font=tfnt, align=la)
        wc(r, 10, n_mismatch,                                fill=tf, font=tfnt, align=ca)
        wc(r, 11, f'{mis_rate:.1f}% of commented returns',   fill=tf, font=sfnt, align=la)

        ws.row_dimensions[r].height = 14
        r += 1
        return r

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = Workbook()

    summary_cols = [
        'ASIN', 'Product Name', 'Total Cases (Returns + Refunds)', 'Physical Returns',
        'Returns with Comment',
        'Match', 'Mismatch', 'Unclear', 'No Comment', 'Refund w/o Return',
        'Mismatch %',
        'Top True Reason (Mismatch)',
        '2nd True Reason (Mismatch)',
        '3rd True Reason (Mismatch)',
    ]

    ws_sum = wb.active
    ws_sum.title = 'SUMMARY'
    ws_sum.append(summary_cols)
    style_header_row(ws_sum, 1, len(summary_cols), C_SUMMARY_H)

    for asin, grp in df.groupby('asin', sort=False):
        pname       = grp['product-name'].iloc[0]
        pname_short = pname[:80] + '...' if len(pname) > 80 else pname
        n_total     = len(grp)
        n_refonly   = (grp['status'] == STATUS_REFUND_ONLY).sum()
        n_returns   = n_total - n_refonly
        n_noc       = (grp['status'] == 'No Comment').sum()
        n_comment   = n_returns - n_noc
        n_match     = (grp['status'] == 'Match').sum()
        n_mis       = (grp['status'] == 'Mismatch').sum()
        n_unc       = (grp['status'] == 'Unclear').sum()
        mis_pct     = round(n_mis / n_comment * 100, 1) if n_comment > 0 else 0

        top_reasons = (
            grp[grp['status'] == 'Mismatch']['true_reason']
            .value_counts().head(3).index.tolist()
        )
        top_reasons += [''] * (3 - len(top_reasons))

        ws_sum.append([
            asin, pname_short, n_total, n_returns, n_comment,
            n_match, n_mis, n_unc, n_noc, n_refonly, mis_pct,
            top_reasons[0], top_reasons[1], top_reasons[2],
        ])

    MIS_PCT_COL, MIS_CNT_COL = 11, 7

    for row_idx in range(2, ws_sum.max_row + 1):
        mis_pct  = ws_sum.cell(row_idx, MIS_PCT_COL).value or 0
        row_fill = PatternFill('solid', fgColor='FFE0E0' if mis_pct >= 20 else
                               ('FFF3CC' if mis_pct >= 10 else 'F0F7F0'))
        for c in range(1, len(summary_cols) + 1):
            cell = ws_sum.cell(row_idx, c)
            cell.font   = body_font
            cell.border = border
            cell.alignment = center_align if c != 2 else left_align
            if c == MIS_PCT_COL:
                cell.number_format = '0.0"%"'
            cell.fill = (
                PatternFill('solid', fgColor='FF9999') if c == MIS_CNT_COL and mis_pct >= 20 else
                PatternFill('solid', fgColor=C_UNCLEAR) if c == MIS_CNT_COL and mis_pct >= 10 else
                PatternFill('solid', fgColor=C_REFUNDONLY) if c == 10 else
                row_fill
            )

    for i, w in enumerate([14, 52, 16, 14, 16, 8, 10, 8, 12, 15, 11, 24, 24, 24], 1):
        ws_sum.column_dimensions[get_column_letter(i)].width = w
    ws_sum.freeze_panes = 'A2'

    DETAIL_COLS = [
        'return-date', 'asin', 'product-name', 'quantity',
        'reason', 'customer-comments',
        'status', 'true_reason', 'comment_topic',
    ]
    DISPLAY_HEADERS = [
        'Return Date', 'ASIN', 'Product Name', 'Qty',
        'Return Reason (Amazon)', 'Customer Comment',
        'Analysis Status', 'True Reason', 'Comment Topic',
    ]

    asin_order = (
        df.groupby('asin')['asin'].count()
        .sort_values(ascending=False).index.tolist()
    )

    for idx, asin in enumerate(asin_order):
        grp   = df[df['asin'] == asin].sort_values('return-date', ascending=False).reset_index(drop=True)
        ws    = wb.create_sheet(title=asin[:31])
        pname = grp['product-name'].iloc[0]

        n_total   = len(grp)
        n_refonly = (grp['status'] == STATUS_REFUND_ONLY).sum()
        n_returns = n_total - n_refonly
        n_noc     = (grp['status'] == 'No Comment').sum()
        n_comment = n_returns - n_noc
        n_match   = (grp['status'] == 'Match').sum()
        n_mis     = (grp['status'] == 'Mismatch').sum()
        n_unc     = (grp['status'] == 'Unclear').sum()
        mis_pct   = round(n_mis / n_comment * 100, 1) if n_comment > 0 else 0

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
        tc = ws.cell(1, 1)
        tc.value     = f"{asin}  |  {pname[:120]}"
        tc.font      = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        tc.fill      = PatternFill('solid', fgColor=C_HEADER)
        tc.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[1].height = 22

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)
        sc = ws.cell(2, 1)
        sc.value = (
            f"Total Cases: {n_total}   |   Physical Returns: {n_returns}   |   "
            f"Refund w/o Return: {n_refonly}   |   With Comment: {n_comment}   |   "
            f"Match: {n_match}   |   Mismatch: {n_mis} ({mis_pct}%)   |   "
            f"Unclear: {n_unc}   |   No Comment: {n_noc}"
        )
        sc.font      = Font(name='Calibri', bold=True, size=9, color='1F3864')
        sc.fill      = PatternFill('solid', fgColor='DCE6F1')
        sc.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[2].height = 15

        analysis_end = write_analysis_block(ws, grp, start_row=3)

        header_row = analysis_end + 1
        data_start = header_row + 1

        for ci, hdr in enumerate(DISPLAY_HEADERS, 1):
            ws.cell(header_row, ci).value = hdr
        style_header_row(ws, header_row, len(DISPLAY_HEADERS))
        ws.freeze_panes = ws.cell(data_start, 1).coordinate

        alt = False
        for row_num, (_, row) in enumerate(grp.iterrows(), data_start):
            alt = not alt
            values = [
                row['return-date'],
                row['asin'],
                row['product-name'],
                row['quantity'],
                row['reason'],
                row['customer-comments'] if not pd.isna(row['customer-comments']) else '',
                row['status'],
                row['true_reason'],
                row['comment_topic'],
            ]
            status = row['status']
            for c_idx, val in enumerate(values, 1):
                cell            = ws.cell(row_num, c_idx)
                cell.value      = val
                cell.font       = body_font
                cell.border     = border
                cell.alignment  = wrap_align if c_idx == 6 else left_align
                if c_idx in (7, 8):
                    cell.fill = STATUS_FILL.get(status, PatternFill('solid', fgColor='FFFFFF'))
                elif c_idx == 5 and status == 'Mismatch':
                    cell.fill = PatternFill('solid', fgColor='FFB3B3')
                else:
                    cell.fill = PatternFill('solid', fgColor=C_ALT_ROW if alt else 'FFFFFF')
            ws.row_dimensions[row_num].height = 14

        for ci, w in enumerate([20, 12, 50, 5, 25, 55, 14, 24, 25, 10, 25], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        for r in range(data_start, ws.max_row + 1):
            ws.cell(r, 1).number_format = 'YYYY-MM-DD HH:MM'

    # ── Save to memory buffer & offer download ────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

# ── 4. Results & download ─────────────────────────────────────────────────────
n_mis   = (df['status'] == 'Mismatch').sum()
n_clear = df['status'].isin(['Match', 'Mismatch']).sum()
mis_pct = round(n_mis / n_clear * 100, 1) if n_clear else 0

st.success("Аналіз завершено!")

n_refonly_total = (df['status'] == STATUS_REFUND_ONLY).sum()

col1, col2, col3, col4 = st.columns(4)
col1.metric("Всього кейсів", len(df))
col2.metric("Фізичні повернення", len(df) - n_refonly_total)
col3.metric("Mismatch", n_mis)
col4.metric("Mismatch %", f"{mis_pct}%")

if n_refund_added:
    st.info(
        f"Додано **{n_refund_added}** refund-ів без фізичного повернення "
        "(returnless refunds, товар не відправлено назад, або повернення ще в дорозі — "
        "вікно повернення ~30 днів). Коментарів у них немає за визначенням, "
        "тому mismatch-аналіз рахується лише по фізичних поверненнях."
    )
if skipped_skus:
    st.caption(f"SKU без жодного повернення (пропущені при мапінгу на ASIN): {skipped_skus}")

st.download_button(
    label="⬇ Завантажити результат (.xlsx)",
    data=buffer,
    file_name="Returns_Analysis.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    type="primary",
)
